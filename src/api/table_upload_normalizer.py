from __future__ import annotations

import csv
import json
import re
import unicodedata
from dataclasses import dataclass
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Literal

from .upload_errors import UploadValidationError


TableFileType = Literal["json", "csv", "spreadsheet"]

TEXT_ENCODINGS: tuple[str, ...] = (
    "utf-8-sig",
    "utf-8",
    "gb18030",
    "big5",
    "shift_jis",
    "cp932",
)
SUMMARY_COLUMN_LIMIT = 50
SUMMARY_NORMALIZATION_LIMIT = 50
SUMMARY_SHEET_LIMIT = 20
SUMMARY_SHEET_COLUMN_LIMIT = 50
MAX_EXCEL_SCAN_ROWS = 10_000
MAX_EXCEL_SCAN_COLUMNS = 500
CSV_DELIMITERS = ",\t;"
XLSX_MAGIC = b"PK"
XLS_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"

_ZERO_WIDTH_AND_FORMAT = {
    "\ufeff",
    "\u200b",
    "\u200c",
    "\u200d",
    "\u2060",
}
_SAFE_FILENAME_CHARS = re.compile(r"[^A-Za-z0-9._-]+")


@dataclass(slots=True, frozen=True)
class SheetSummary:
    sheet_name: str
    columns: list[str]
    row_count: int
    column_count: int
    columns_truncated: bool = False

    def to_preview(self) -> dict[str, Any]:
        return {
            "sheet_name": self.sheet_name,
            "columns": list(self.columns),
            "row_count": self.row_count,
            "column_count": self.column_count,
            "columns_truncated": self.columns_truncated,
        }


@dataclass(slots=True, frozen=True)
class TableNormalizationResult:
    file_type: TableFileType
    normalized_content_text: str | None
    normalized_content_type: str | None
    normalized_filename: str | None
    preview: dict[str, Any]
    requires_sheet_selection: bool = False
    selected_sheet: str | None = None
    excel_sheets: tuple[SheetSummary, ...] = ()


@dataclass(slots=True, frozen=True)
class _JsonObjectPairs:
    pairs: list[tuple[str, Any]]


def detect_table_file_type(filename: str, content_type: str | None, content: bytes) -> TableFileType | None:
    suffix = Path(filename).suffix.lower()
    if suffix == ".json":
        return "json"
    if suffix == ".csv":
        return "csv"
    if suffix in {".xlsx", ".xls"}:
        return "spreadsheet"
    if _looks_like_excel(content):
        return "spreadsheet"
    base_content_type = (content_type or "").split(";", 1)[0].strip().lower()
    if base_content_type in {"application/json", "text/json"}:
        return "json"
    if base_content_type in {"text/csv", "application/csv", "application/vnd.ms-excel"}:
        return "csv"
    return None


def normalize_table_upload(
    *,
    filename: str,
    content_type: str | None,
    content: bytes,
    selected_sheet: str | None = None,
) -> TableNormalizationResult:
    file_type = detect_table_file_type(filename, content_type, content)
    if file_type is None:
        raise UploadValidationError("Only JSON, CSV, Excel, PNG, JPG/JPEG, and PDF files are supported")
    if file_type == "csv":
        decoded = _decode_text(content)
        return _normalize_csv(decoded.text, source_encoding=decoded.encoding, filename=filename)
    if file_type == "json":
        decoded = _decode_text(content)
        return _normalize_json(decoded.text, source_encoding=decoded.encoding, filename=filename)
    return _normalize_spreadsheet(filename=filename, content=content, selected_sheet=selected_sheet)


def normalize_selected_spreadsheet_sheet(
    *,
    filename: str,
    content_type: str | None,
    content: bytes,
    selected_sheet: str,
) -> TableNormalizationResult:
    result = normalize_table_upload(
        filename=filename,
        content_type=content_type,
        content=content,
        selected_sheet=selected_sheet,
    )
    if result.normalized_content_text is None:
        raise UploadValidationError("Selected spreadsheet sheet did not produce executable content")
    return result


@dataclass(slots=True, frozen=True)
class _DecodedText:
    text: str
    encoding: str


def _decode_text(content: bytes) -> _DecodedText:
    for encoding in TEXT_ENCODINGS:
        try:
            return _DecodedText(content.decode(encoding, errors="strict"), encoding)
        except UnicodeDecodeError:
            continue
    raise UploadValidationError("Unable to detect text encoding; please save as UTF-8 CSV/JSON or Excel and upload again")


def _normalize_csv(text: str, *, source_encoding: str, filename: str) -> TableNormalizationResult:
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=CSV_DELIMITERS)
    except csv.Error:
        dialect = csv.excel
    try:
        rows = list(csv.reader(StringIO(text), dialect=dialect))
    except csv.Error as exc:
        raise UploadValidationError(f"Invalid CSV file: {exc}") from exc
    if not rows:
        raise UploadValidationError("CSV upload must include a header row")
    original_columns = ["" if value is None else str(value) for value in rows[0]]
    cleaned_columns, column_normalizations = _clean_columns(original_columns)
    output = StringIO()
    writer = csv.writer(output, lineterminator="\n")
    writer.writerow(cleaned_columns)
    writer.writerows(rows[1:])
    normalized_text = output.getvalue()
    preview = _table_preview(
        row_count=len(rows) - 1,
        columns=cleaned_columns,
        shape="table",
        source_encoding=source_encoding,
        original_columns=original_columns,
        column_normalizations=column_normalizations,
        extra={"normalized_content_type": "text/csv"},
    )
    return TableNormalizationResult(
        file_type="csv",
        normalized_content_text=normalized_text,
        normalized_content_type="text/csv",
        normalized_filename=_normalized_csv_filename(filename),
        preview=preview,
    )


def _normalize_json(text: str, *, source_encoding: str, filename: str) -> TableNormalizationResult:
    try:
        value = json.loads(text, object_pairs_hook=lambda pairs: _JsonObjectPairs(list(pairs)))
    except json.JSONDecodeError as exc:
        raise UploadValidationError(f"Invalid JSON file: {exc}") from exc
    if isinstance(value, _JsonObjectPairs):
        normalized, columns, normalizations = _normalize_json_object_pairs(value)
        row_count = 1
        shape = "object"
    elif isinstance(value, list):
        normalized_rows: list[Any] = []
        columns: list[str] = []
        normalizations: list[dict[str, Any]] = []
        for row in value:
            if isinstance(row, _JsonObjectPairs):
                normalized_row, row_columns, row_normalizations = _normalize_json_object_pairs(row)
                normalized_rows.append(normalized_row)
                for column in row_columns:
                    if column not in columns:
                        columns.append(column)
                normalizations.extend(row_normalizations)
            else:
                normalized_rows.append(_json_plain(row))
        normalized = normalized_rows
        row_count = len(value)
        shape = "array"
    else:
        raise UploadValidationError("JSON upload must be an object or an array")
    normalized_text = json.dumps(normalized, ensure_ascii=False, separators=(",", ":"))
    preview = _table_preview(
        row_count=row_count,
        columns=columns,
        shape=shape,
        source_encoding=source_encoding,
        original_columns=[item["original"] for item in normalizations],
        column_normalizations=normalizations,
        extra={"normalized_content_type": "application/json"},
    )
    return TableNormalizationResult(
        file_type="json",
        normalized_content_text=normalized_text,
        normalized_content_type="application/json",
        normalized_filename=_normalized_json_filename(filename),
        preview=preview,
    )


def _normalize_json_object_pairs(value: _JsonObjectPairs) -> tuple[dict[str, Any], list[str], list[dict[str, Any]]]:
    original_keys = [str(key) for key, _child in value.pairs]
    cleaned_keys, normalizations = _clean_columns(original_keys)
    normalized: dict[str, Any] = {}
    for cleaned_key, (_original_key, child) in zip(cleaned_keys, value.pairs, strict=False):
        normalized[cleaned_key] = _json_plain(child)
    return normalized, cleaned_keys, normalizations


def _json_plain(value: Any) -> Any:
    if isinstance(value, _JsonObjectPairs):
        return {str(key): _json_plain(child) for key, child in value.pairs}
    if isinstance(value, list):
        return [_json_plain(item) for item in value]
    return value


def _normalize_spreadsheet(
    *,
    filename: str,
    content: bytes,
    selected_sheet: str | None,
) -> TableNormalizationResult:
    sheets = _read_spreadsheet_sheets(filename=filename, content=content)
    if not sheets:
        raise UploadValidationError("Spreadsheet upload must include at least one sheet with a header row")
    if len(sheets) > SUMMARY_SHEET_LIMIT:
        raise UploadValidationError(f"Spreadsheet contains more than {SUMMARY_SHEET_LIMIT} valid sheets; please split the file")
    selected = selected_sheet
    if selected is None and len(sheets) == 1:
        selected = sheets[0].sheet_name
    sheet_names = [sheet.sheet_name for sheet in sheets]
    if selected is not None and selected not in sheet_names:
        raise UploadValidationError(f"Unknown spreadsheet sheet selection: {selected}")
    selected_summary = next((sheet for sheet in sheets if sheet.sheet_name == selected), None)
    normalized_text = None
    normalized_filename = None
    if selected_summary is not None:
        normalized_text = _spreadsheet_sheet_to_csv(filename=filename, content=content, sheet_name=selected_summary.sheet_name)
        normalized_filename = _normalized_spreadsheet_filename(filename, selected_summary.sheet_name if len(sheets) > 1 else None)
    preview = {
        "row_count": selected_summary.row_count if selected_summary is not None else None,
        "columns": list(selected_summary.columns) if selected_summary is not None else [],
        "shape": "spreadsheet",
        "file_type": "spreadsheet",
        "normalized_content_type": "text/csv" if selected_summary is not None else None,
        "requires_sheet_selection": len(sheets) > 1 and selected_sheet is None,
        "selected_sheet": selected_summary.sheet_name if selected_summary is not None else None,
        "excel_sheets": [sheet.to_preview() for sheet in sheets],
        "excel_sheet_count": len(sheets),
        "excel_sheets_truncated": False,
    }
    if selected_summary is not None:
        preview.update(
            {
                "column_count": selected_summary.column_count,
                "columns_truncated": selected_summary.columns_truncated,
            }
        )
    return TableNormalizationResult(
        file_type="spreadsheet",
        normalized_content_text=normalized_text,
        normalized_content_type="text/csv" if normalized_text is not None else None,
        normalized_filename=normalized_filename,
        preview=preview,
        requires_sheet_selection=len(sheets) > 1 and selected_sheet is None,
        selected_sheet=selected_summary.sheet_name if selected_summary is not None else None,
        excel_sheets=tuple(sheets),
    )


def _read_spreadsheet_sheets(*, filename: str, content: bytes) -> list[SheetSummary]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".xls" and content.startswith(XLS_MAGIC):
        return _read_xls_sheets(content)
    if suffix == ".xlsx" or content.startswith(XLSX_MAGIC):
        return _read_xlsx_sheets(content)
    if content.startswith(XLS_MAGIC):
        return _read_xls_sheets(content)
    raise UploadValidationError("Unsupported or invalid Excel file")


def _read_xlsx_sheets(content: bytes) -> list[SheetSummary]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - environment contract guard.
        raise UploadValidationError("Excel .xlsx support is not installed") from exc
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True, keep_links=False)
    except Exception as exc:  # noqa: BLE001 - parser errors become user-safe validation.
        raise UploadValidationError("Invalid .xlsx spreadsheet") from exc
    try:
        summaries: list[SheetSummary] = []
        for worksheet in workbook.worksheets:
            summary = _sheet_summary_from_rows(
                worksheet.title,
                worksheet.iter_rows(values_only=True),
            )
            if summary is not None:
                summaries.append(summary)
        return summaries
    finally:
        workbook.close()


def _read_xls_sheets(content: bytes) -> list[SheetSummary]:
    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover - environment contract guard.
        raise UploadValidationError("Excel .xls support is not installed") from exc
    try:
        workbook = xlrd.open_workbook(file_contents=content, on_demand=True)
    except Exception as exc:  # noqa: BLE001
        raise UploadValidationError("Invalid .xls spreadsheet") from exc
    summaries: list[SheetSummary] = []
    try:
        for sheet_name in workbook.sheet_names():
            sheet = workbook.sheet_by_name(sheet_name)
            rows = ([sheet.cell_value(rowx, colx) for colx in range(sheet.ncols)] for rowx in range(sheet.nrows))
            summary = _sheet_summary_from_rows(sheet_name, rows)
            if summary is not None:
                summaries.append(summary)
    finally:
        workbook.release_resources()
    return summaries


def _spreadsheet_sheet_to_csv(*, filename: str, content: bytes, sheet_name: str) -> str:
    suffix = Path(filename).suffix.lower()
    if suffix == ".xls" and content.startswith(XLS_MAGIC):
        rows = _xls_sheet_rows(content, sheet_name)
    else:
        rows = _xlsx_sheet_rows(content, sheet_name)
    header_index, original_columns = _first_header_row(rows)
    cleaned_columns, _normalizations = _clean_columns(original_columns)
    output = StringIO()
    writer = csv.writer(output, lineterminator="\n")
    writer.writerow(cleaned_columns)
    for row in rows[header_index + 1 :]:
        if _row_is_empty(row):
            continue
        writer.writerow(["" if value is None else value for value in row[: len(original_columns)]])
    return output.getvalue()


def _xlsx_sheet_rows(content: bytes, sheet_name: str) -> list[list[Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True, keep_links=False)
    try:
        if sheet_name not in workbook.sheetnames:
            raise UploadValidationError(f"Unknown spreadsheet sheet selection: {sheet_name}")
        worksheet = workbook[sheet_name]
        return [list(row) for row in worksheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def _xls_sheet_rows(content: bytes, sheet_name: str) -> list[list[Any]]:
    import xlrd

    workbook = xlrd.open_workbook(file_contents=content, on_demand=True)
    try:
        if sheet_name not in workbook.sheet_names():
            raise UploadValidationError(f"Unknown spreadsheet sheet selection: {sheet_name}")
        sheet = workbook.sheet_by_name(sheet_name)
        return [[sheet.cell_value(rowx, colx) for colx in range(sheet.ncols)] for rowx in range(sheet.nrows)]
    finally:
        workbook.release_resources()


def _sheet_summary_from_rows(sheet_name: str, row_iter: Any) -> SheetSummary | None:
    rows: list[list[Any]] = []
    for index, row in enumerate(row_iter):
        if index >= MAX_EXCEL_SCAN_ROWS:
            break
        row_values = list(row)[:MAX_EXCEL_SCAN_COLUMNS]
        rows.append(row_values)
    try:
        header_index, original_columns = _first_header_row(rows)
    except UploadValidationError:
        return None
    cleaned_columns, _normalizations = _clean_columns(original_columns)
    row_count = sum(1 for row in rows[header_index + 1 :] if not _row_is_empty(row))
    columns_preview, truncated = _truncate(cleaned_columns, SUMMARY_SHEET_COLUMN_LIMIT)
    return SheetSummary(
        sheet_name=str(sheet_name),
        columns=columns_preview,
        row_count=row_count,
        column_count=len(cleaned_columns),
        columns_truncated=truncated,
    )


def _first_header_row(rows: list[list[Any]]) -> tuple[int, list[str]]:
    for index, row in enumerate(rows):
        if _row_is_empty(row):
            continue
        last_non_empty = 0
        for cell_index, value in enumerate(row):
            if not _cell_is_empty(value):
                last_non_empty = cell_index + 1
        return index, ["" if value is None else _cell_to_text(value) for value in row[:last_non_empty]]
    raise UploadValidationError("Spreadsheet sheet must include a header row")


def _cell_to_text(value: Any) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _row_is_empty(row: list[Any]) -> bool:
    return all(_cell_is_empty(value) for value in row)


def _cell_is_empty(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _clean_columns(original_columns: list[str]) -> tuple[list[str], list[dict[str, Any]]]:
    if not original_columns:
        raise UploadValidationError("Table upload must include a header row")
    cleaned = [_clean_header(value) for value in original_columns]
    empty_positions = [index + 1 for index, value in enumerate(cleaned) if not value]
    if empty_positions:
        raise UploadValidationError(f"Table header contains empty normalized column names at positions: {empty_positions}")
    seen: dict[str, int] = {}
    duplicates: list[str] = []
    for index, value in enumerate(cleaned, start=1):
        if value in seen:
            duplicates.append(value)
        seen[value] = index
    if duplicates:
        raise UploadValidationError(f"Table header contains duplicate normalized column names: {', '.join(sorted(set(duplicates)))}")
    normalizations = [
        {
            "index": index,
            "original": original,
            "normalized": normalized,
            "changed": original != normalized,
        }
        for index, (original, normalized) in enumerate(zip(original_columns, cleaned, strict=False), start=1)
    ]
    return cleaned, normalizations


def _clean_header(value: Any) -> str:
    text = unicodedata.normalize("NFKC", "" if value is None else str(value))
    text = "".join(_clean_header_char(char) for char in text)
    text = text.strip()
    previous = None
    while previous != text:
        previous = text
        text = _strip_paired_quotes(text.strip())
    return text.strip()


def _clean_header_char(char: str) -> str:
    if char in _ZERO_WIDTH_AND_FORMAT:
        return ""
    category = unicodedata.category(char)
    if category in {"Cf", "Cc"}:
        return ""
    return char


def _strip_paired_quotes(text: str) -> str:
    quote_pairs = {
        '"': '"',
        "'": "'",
        "“": "”",
        "‘": "’",
        "「": "」",
        "『": "』",
    }
    if len(text) >= 2 and quote_pairs.get(text[0]) == text[-1]:
        return text[1:-1]
    return text


def _table_preview(
    *,
    row_count: int,
    columns: list[str],
    shape: str,
    source_encoding: str,
    original_columns: list[str],
    column_normalizations: list[dict[str, Any]],
    extra: dict[str, Any] | None = None,
) -> dict[str, Any]:
    columns_preview, columns_truncated = _truncate(columns, SUMMARY_COLUMN_LIMIT)
    normalizations_preview, normalizations_truncated = _truncate(column_normalizations, SUMMARY_NORMALIZATION_LIMIT)
    return {
        "row_count": row_count,
        "columns": columns_preview,
        "shape": shape,
        "source_encoding": source_encoding,
        "original_columns": _truncate(original_columns, SUMMARY_COLUMN_LIMIT)[0],
        "column_normalizations": normalizations_preview,
        "column_count": len(columns),
        "columns_truncated": columns_truncated,
        "column_normalization_count": len(column_normalizations),
        "column_normalizations_truncated": normalizations_truncated,
        **(extra or {}),
    }


def _truncate(values: list[Any], limit: int) -> tuple[list[Any], bool]:
    return list(values[:limit]), len(values) > limit


def _normalized_csv_filename(filename: str) -> str:
    suffix = Path(filename).suffix.lower()
    return _safe_filename(filename if suffix == ".csv" else f"{Path(filename).stem or 'upload'}.csv")


def _normalized_json_filename(filename: str) -> str:
    suffix = Path(filename).suffix.lower()
    return _safe_filename(filename if suffix == ".json" else f"{Path(filename).stem or 'upload'}.json")


def _normalized_spreadsheet_filename(filename: str, sheet_name: str | None) -> str:
    stem = Path(filename).stem or "spreadsheet"
    if sheet_name:
        stem = f"{stem}-{sheet_name}"
    return _safe_filename(f"{stem}.csv")


def _safe_filename(filename: str) -> str:
    stem = Path(filename).stem.strip() or "upload"
    suffix = Path(filename).suffix.lower() or ".csv"
    safe_stem = _SAFE_FILENAME_CHARS.sub("_", unicodedata.normalize("NFKC", stem)).strip("._-")
    return f"{safe_stem or 'upload'}{suffix}"


def _looks_like_excel(content: bytes) -> bool:
    return content.startswith(XLS_MAGIC) or content.startswith(XLSX_MAGIC)
