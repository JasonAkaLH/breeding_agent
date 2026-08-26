from __future__ import annotations

import base64
import csv
from dataclasses import replace
from datetime import datetime
import gzip
from io import BytesIO, StringIO
import json
import unittest
from dataclasses import FrozenInstanceError
from unittest.mock import AsyncMock, patch

from openpyxl import Workbook

from src.storage.conversation_files import FILE_UPLOAD_MESSAGE_TYPE, file_upload_message_id
from src.api.routes.uploads import _read_upload_content_with_limit
from src.api.upload_store import DEFAULT_MAX_UPLOAD_FILE_BYTES, InMemoryUploadStore, UploadValidationError
from src.orchestration.agent_loop.orchestrator import AgentExecutionRequest
from tests.api.support import APITestCase


class _ChunkedUpload:
    def __init__(self, chunks):
        self._chunks = list(chunks)
        self.read_calls = 0

    async def read(self, size: int = -1):
        self.read_calls += 1
        if not self._chunks:
            return b""
        return self._chunks.pop(0)


class UploadReadLimitTest(unittest.IsolatedAsyncioTestCase):
    def test_default_upload_limit_is_twenty_mebibytes(self) -> None:
        self.assertEqual(DEFAULT_MAX_UPLOAD_FILE_BYTES, 20 * 1024 * 1024)
        self.assertEqual(InMemoryUploadStore().max_file_bytes, DEFAULT_MAX_UPLOAD_FILE_BYTES)

    async def test_read_upload_content_stops_as_soon_as_limit_is_exceeded(self) -> None:
        upload = _ChunkedUpload([b"12345", b"67890", b"extra-should-not-be-read"])

        with self.assertRaisesRegex(UploadValidationError, "exceeds 9 bytes"):
            await _read_upload_content_with_limit(upload, max_bytes=9, chunk_size=5)

        self.assertEqual(upload.read_calls, 2)

    async def test_read_upload_content_allows_exact_limit(self) -> None:
        upload = _ChunkedUpload([b"12345", b"67890"])

        content = await _read_upload_content_with_limit(upload, max_bytes=10, chunk_size=5)

        self.assertEqual(content, b"1234567890")
        self.assertEqual(upload.read_calls, 3)

    def test_upload_store_rejects_pathful_filename_instead_of_silent_normalization(self) -> None:
        store = InMemoryUploadStore()

        with self.assertRaisesRegex(UploadValidationError, "filename"):
            store.save(
                username="acc-1",
                conversation_id="conv-1",
                filename="../secret.csv",
                content_type="text/csv",
                content=b"col\nvalue\n",
            )


class UploadsAPITest(APITestCase):
    async def asyncSetUp(self) -> None:
        await super().asyncSetUp()
        await self.reconfigure_runtime(main_agent_stream_generator=lambda _prompt: ["已收到。"])

    async def test_upload_csv_returns_preview_and_submit_resolves_for_skill(self) -> None:
        csv_content = "ped_id,design_check,set\nCK_A,1,A\nA001,0,A\n"

        upload = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-upload"},
            files={"file": ("rcbd.csv", csv_content, "text/csv")},
        )

        self.assertEqual(upload.status_code, 201)
        payload = upload.json()
        self.assertEqual(payload["filename"], "rcbd.csv")
        self.assertEqual(payload["file_type"], "csv")
        self.assertEqual(payload["preview"]["columns"], ["ped_id", "design_check", "set"])
        self.assertEqual(payload["preview"]["row_count"], 2)
        self.assertNotIn("content", payload)

        resolved = await self.runtime.resolve_uploads_for_message(
            "conv-upload", "acc-1", [payload["upload_id"]]
        )
        self.assertEqual(resolved["uploaded_artifacts"][0]["filename"], "rcbd.csv")
        self.assertNotIn("content", resolved["uploaded_artifacts"][0])
        self.assertEqual(resolved["skill_artifacts"][0]["content"], csv_content)

        submitted = await self.submit_message(
            conversation_id="conv-upload",
            content="请用这个CSV做3个区组的RCBD设计",
            capability_id=None,
            metadata={"upload_ids": [payload["upload_id"]], "blocks": 3},
        )
        self.assertEqual(submitted.status_code, 202)
        task_id = submitted.json()["task_id"]
        await self.wait_for_terminal_task(task_id)

    async def test_upload_success_writes_file_upload_history_message(self) -> None:
        upload = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-upload-history"},
            files={"file": ("materials.csv", "ped_id,value\nA001,1\n", "text/csv")},
        )

        self.assertEqual(upload.status_code, 201, upload.text)
        upload_id = upload.json()["upload_id"]
        response = await self.client.get("/api/v1/conversations/conv-upload-history/messages")

        self.assertEqual(response.status_code, 200, response.text)
        file_messages = [
            message
            for message in response.json()["messages"]
            if message["message_id"] == file_upload_message_id(upload_id)
        ]
        self.assertEqual(len(file_messages), 1)
        file_message = file_messages[0]
        self.assertEqual(file_message["role"], "system")
        self.assertEqual(file_message["message_type"], FILE_UPLOAD_MESSAGE_TYPE)
        self.assertEqual(file_message["metadata"]["upload_id"], upload_id)
        self.assertEqual(file_message["metadata"]["filename"], "materials.csv")
        self.assertEqual(file_message["metadata"]["file_status"], "active")
        self.assertEqual(file_message["metadata"]["description_status"], "ready")
        serialized = json.dumps(file_message, ensure_ascii=False)
        self.assertNotIn("storage_key", serialized)
        self.assertNotIn("content_base64", serialized)
        self.assertTrue((self.runtime.conversation_file_store.conversation_dir("conv-upload-history") / "index.md").exists())

    async def test_upload_index_write_transient_failure_retries_and_succeeds_without_marker(self) -> None:
        original_write_index = self.runtime._conversation_file_index_writer.write_index
        attempts = 0

        def flaky_write_index(*args, **kwargs):
            nonlocal attempts
            attempts += 1
            if attempts == 1:
                raise OSError("transient index failure")
            return original_write_index(*args, **kwargs)

        self.runtime._conversation_file_index_writer.write_index = flaky_write_index
        try:
            upload = await self.client.post(
                "/api/v1/conversations/uploads",
                data={"conversation_id": "conv-upload-transient-index"},
                files={"file": ("materials.csv", "ped_id,value\nA001,1\n", "text/csv")},
            )
        finally:
            self.runtime._conversation_file_index_writer.write_index = original_write_index

        self.assertEqual(upload.status_code, 201, upload.text)
        marker = await self.runtime.storage.get_conversation_file_index_repair_marker("conv-upload-transient-index")
        self.assertIsNone(marker)
        self.assertEqual(attempts, 2)

    async def test_upload_index_write_failure_fails_closed_and_records_repair_marker(self) -> None:
        await self.runtime.storage.record_conversation_file_index_repair_required(
            "conv-upload-index-fail",
            reason_code="preexisting_index_write_failed",
            affected_upload_ids=("upl-existing",),
            now=self.runtime._utcnow_naive(),
        )
        original_write_index = self.runtime._conversation_file_index_writer.write_index

        def failing_write_index(*_args, **_kwargs):
            raise OSError("index unavailable")

        self.runtime._conversation_file_index_writer.write_index = failing_write_index
        try:
            upload = await self.client.post(
                "/api/v1/conversations/uploads",
                data={"conversation_id": "conv-upload-index-fail"},
                files={"file": ("materials.csv", "ped_id,value\nA001,1\n", "text/csv")},
            )
        finally:
            self.runtime._conversation_file_index_writer.write_index = original_write_index

        self.assertEqual(upload.status_code, 400, upload.text)
        self.assertIn("indexed", upload.json()["detail"])
        marker = await self.runtime.storage.get_conversation_file_index_repair_marker("conv-upload-index-fail")
        self.assertIsNotNone(marker)
        self.assertEqual(marker.status, "pending")
        self.assertEqual(marker.reason_code, "upload_index_write_failed")
        self.assertEqual(marker.affected_upload_ids[0], "upl-existing")
        upload_id = next(upload_id for upload_id in marker.affected_upload_ids if upload_id != "upl-existing")
        self.assertIsNone(
            await self.runtime.storage.get_conversation_file_resource(
                "conv-upload-index-fail",
                "acc-1",
                upload_id,
            )
        )
        self.assertIsNone(await self.runtime.storage.get_message(file_upload_message_id(upload_id)))
        with self.assertRaises(UploadValidationError):
            self.runtime.upload_store.get_for_message(
                upload_id=upload_id,
                username="acc-1",
                conversation_id="conv-upload-index-fail",
            )
        self.assertFalse(
            (self.runtime.conversation_file_store.conversation_dir("conv-upload-index-fail") / upload_id).exists()
        )
        audit_log = (self.workspace / "audit.jsonl").read_text(encoding="utf-8")
        self.assertIn("conversation_file.file_upload_index_repair_required", audit_log)
        self.assertIn("upload_index_write_failed", audit_log)
        self.assertNotIn("storage_key", audit_log)
        self.assertNotIn("content_base64", audit_log)
        self.assertNotIn("ped_id,value", audit_log)

    async def test_upload_composite_db_failure_compensates_local_and_memory_state(self) -> None:
        original_save_composite = self.runtime.storage.save_conversation_file_resource_with_upload_message

        async def failing_save_composite(*_args, **_kwargs):
            raise RuntimeError("composite db write failed")

        self.runtime.storage.save_conversation_file_resource_with_upload_message = failing_save_composite
        try:
            upload = await self.client.post(
                "/api/v1/conversations/uploads",
                data={"conversation_id": "conv-upload-db-fail"},
                files={"file": ("materials.csv", "ped_id,value\nA001,1\n", "text/csv")},
            )
        finally:
            self.runtime.storage.save_conversation_file_resource_with_upload_message = original_save_composite

        self.assertEqual(upload.status_code, 400, upload.text)
        resources = await self.runtime.storage.list_conversation_file_resources(
            "conv-upload-db-fail",
            "acc-1",
            include_deleted=True,
        )
        messages = await self.runtime.storage.list_messages_for_conversation("conv-upload-db-fail")
        memory_records = self.runtime.upload_store.list_for_conversation(
            username="acc-1",
            conversation_id="conv-upload-db-fail",
        )
        conversation_dir = self.runtime.conversation_file_store.conversation_dir("conv-upload-db-fail")
        child_dirs = [child for child in conversation_dir.iterdir()] if conversation_dir.exists() else []
        self.assertEqual(resources, [])
        self.assertEqual(messages, [])
        self.assertEqual(memory_records, [])
        self.assertEqual(child_dirs, [])

    async def test_upload_description_write_failure_compensates_local_and_memory_state(self) -> None:
        original_write_description = self.runtime.conversation_file_store.write_description

        def failing_write_description(*_args, **_kwargs):
            raise OSError("description write failed")

        self.runtime.conversation_file_store.write_description = failing_write_description
        try:
            upload = await self.client.post(
                "/api/v1/conversations/uploads",
                data={"conversation_id": "conv-upload-description-fail"},
                files={"file": ("materials.csv", "ped_id,value\nA001,1\n", "text/csv")},
            )
        finally:
            self.runtime.conversation_file_store.write_description = original_write_description

        self.assertEqual(upload.status_code, 400, upload.text)
        resources = await self.runtime.storage.list_conversation_file_resources(
            "conv-upload-description-fail",
            "acc-1",
            include_deleted=True,
        )
        memory_records = self.runtime.upload_store.list_for_conversation(
            username="acc-1",
            conversation_id="conv-upload-description-fail",
        )
        conversation_dir = self.runtime.conversation_file_store.conversation_dir("conv-upload-description-fail")
        child_dirs = [child for child in conversation_dir.iterdir()] if conversation_dir.exists() else []
        self.assertEqual(resources, [])
        self.assertEqual(memory_records, [])
        self.assertEqual(child_dirs, [])

    async def test_upload_index_failure_marker_write_failure_still_fails_closed(self) -> None:
        original_write_index = self.runtime._conversation_file_index_writer.write_index
        original_record_marker = self.runtime.storage.record_conversation_file_index_repair_required

        def failing_write_index(*_args, **_kwargs):
            raise OSError("index unavailable")

        async def failing_record_marker(*_args, **_kwargs):
            raise RuntimeError("marker store unavailable")

        self.runtime._conversation_file_index_writer.write_index = failing_write_index
        self.runtime.storage.record_conversation_file_index_repair_required = failing_record_marker
        try:
            upload = await self.client.post(
                "/api/v1/conversations/uploads",
                data={"conversation_id": "conv-upload-marker-fail"},
                files={"file": ("materials.csv", "ped_id,value\nA001,1\n", "text/csv")},
            )
        finally:
            self.runtime._conversation_file_index_writer.write_index = original_write_index
            self.runtime.storage.record_conversation_file_index_repair_required = original_record_marker

        self.assertEqual(upload.status_code, 400, upload.text)
        resources = await self.runtime.storage.list_conversation_file_resources(
            "conv-upload-marker-fail",
            "acc-1",
            include_deleted=True,
        )
        messages = await self.runtime.storage.list_messages_for_conversation("conv-upload-marker-fail")
        marker = await self.runtime.storage.get_conversation_file_index_repair_marker("conv-upload-marker-fail")
        self.assertEqual(resources, [])
        self.assertEqual(messages, [])
        self.assertIsNone(marker)

    async def test_delete_upload_marks_history_deleted_and_rewrites_index(self) -> None:
        upload = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-delete-history"},
            files={"file": ("delete-me.csv", "ped_id,value\nA001,1\n", "text/csv")},
        )
        self.assertEqual(upload.status_code, 201, upload.text)
        upload_id = upload.json()["upload_id"]

        deleted = await self.client.request(
            "DELETE",
            "/api/v1/conversations/uploads",
            json={"conversation_id": "conv-delete-history", "upload_id": upload_id},
        )

        self.assertEqual(deleted.status_code, 200, deleted.text)
        self.assertTrue(deleted.json()["deleted"])
        resource = await self.runtime.storage.get_conversation_file_resource(
            "conv-delete-history",
            "acc-1",
            upload_id,
        )
        message = await self.runtime.storage.get_message(file_upload_message_id(upload_id))
        self.assertEqual(resource.status, "deleted")
        self.assertEqual(message.metadata["file_status"], "deleted")
        index_text = (self.runtime.conversation_file_store.conversation_dir("conv-delete-history") / "index.md").read_text()
        self.assertIn("文件本体已物理删除", index_text)

    async def test_deleted_resource_not_in_conversation_upload_context(self) -> None:
        active = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-deleted-context"},
            files={"file": ("active.csv", "ped_id,value\nA001,1\n", "text/csv")},
        )
        deleted_upload = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-deleted-context"},
            files={"file": ("deleted.csv", "ped_id,value\nB001,2\n", "text/csv")},
        )
        self.assertEqual(active.status_code, 201, active.text)
        self.assertEqual(deleted_upload.status_code, 201, deleted_upload.text)
        active_id = active.json()["upload_id"]
        deleted_id = deleted_upload.json()["upload_id"]
        deleted = await self.client.request(
            "DELETE",
            "/api/v1/conversations/uploads",
            json={"conversation_id": "conv-deleted-context", "upload_id": deleted_id},
        )
        self.assertEqual(deleted.status_code, 200, deleted.text)

        context = await self.runtime.resolve_conversation_uploads_for_message("conv-deleted-context", "acc-1")

        serialized = json.dumps(context, ensure_ascii=False, default=str)
        self.assertIn(active_id, serialized)
        self.assertNotIn(deleted_id, serialized)
        self.assertNotIn("deleted.csv", serialized)

    async def test_default_context_after_delete_scrubs_stale_upload_metadata(self) -> None:
        upload = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-stale-metadata"},
            files={"file": ("stale.csv", "ped_id,value\nA001,1\n", "text/csv")},
        )
        self.assertEqual(upload.status_code, 201, upload.text)
        upload_id = upload.json()["upload_id"]
        context = await self.runtime.resolve_conversation_uploads_for_message("conv-stale-metadata", "acc-1")
        deleted = await self.client.request(
            "DELETE",
            "/api/v1/conversations/uploads",
            json={"conversation_id": "conv-stale-metadata", "upload_id": upload_id},
        )
        self.assertEqual(deleted.status_code, 200, deleted.text)

        request = AgentExecutionRequest(
            task_id="task-stale",
            conversation_id="conv-stale-metadata",
            root_message_id="msg-stale",
            user_message="use stale file",
            owner_scope="owner:test",
            metadata={
                "uploaded_artifacts": context["uploaded_artifacts"],
                "skill_artifacts": context["skill_artifacts"],
            },
        )
        scrubbed = await self.runtime._scrub_deleted_file_context_for_execution(request)

        serialized = json.dumps(scrubbed.metadata, ensure_ascii=False, default=str)
        self.assertNotIn(upload_id, serialized)
        self.assertNotIn("stale.csv", serialized)
        self.assertNotIn("storage_key", serialized)

    async def test_task_bound_upload_deleted_before_execution_fails_closed(self) -> None:
        upload = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-bound-delete-before-execution"},
            files={"file": ("bound.csv", "ped_id,value\nA001,1\n", "text/csv")},
        )
        self.assertEqual(upload.status_code, 201, upload.text)
        upload_id = upload.json()["upload_id"]
        initialized_runs: list[object] = []
        main_agent_called = False

        async def capture_execution(initialized: object) -> None:
            initialized_runs.append(initialized)

        def fail_if_called(_prompt: str, **_kwargs):
            nonlocal main_agent_called
            main_agent_called = True
            return "should not run"

        await self.reconfigure_runtime(main_agent_stream_generator=fail_if_called)
        self.runtime._schedule_initialized_execution = capture_execution
        response = await self.submit_message(
            conversation_id="conv-bound-delete-before-execution",
            capability_id=None,
            content="请使用显式绑定的文件。",
            metadata={"upload_ids": [upload_id]},
        )
        self.assertEqual(response.status_code, 202, response.text)
        self.assertEqual(len(initialized_runs), 1)
        task_id = response.json()["task_id"]
        attachments = await self.runtime.storage.list_task_input_attachments_for_task(task_id)
        self.assertEqual([attachment.source_upload_id for attachment in attachments], [upload_id])

        deleted = await self.client.request(
            "DELETE",
            "/api/v1/conversations/uploads",
            json={"conversation_id": "conv-bound-delete-before-execution", "upload_id": upload_id},
        )
        self.assertEqual(deleted.status_code, 200, deleted.text)

        await self.runtime._run_initialized_execution(
            initialized_runs[0],
            execution_generation=1,
        )

        task = await self.runtime.storage.get_task(task_id)
        self.assertEqual(str(task.status), "failed")
        self.assertFalse(main_agent_called)
        events = await self.runtime.storage.list_events_for_task(task_id)
        failed = next(event for event in events if event.event_type == "task.failed")
        self.assertEqual(failed.payload["code"], "execution_crash")
        self.assertEqual(failed.payload["message"], "Task execution failed safely.")
        self.assertNotIn(upload_id, failed.payload["message"])

    async def test_delete_index_failure_keeps_deleted_fact_and_records_repair_marker(self) -> None:
        upload = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-delete-index-fail"},
            files={"file": ("delete-me.csv", "ped_id,value\nA001,1\n", "text/csv")},
        )
        self.assertEqual(upload.status_code, 201, upload.text)
        upload_id = upload.json()["upload_id"]
        original_write_index = self.runtime._conversation_file_index_writer.write_index

        def failing_write_index(*_args, **_kwargs):
            raise OSError("index unavailable")

        self.runtime._conversation_file_index_writer.write_index = failing_write_index
        try:
            deleted = await self.client.request(
                "DELETE",
                "/api/v1/conversations/uploads",
                json={"conversation_id": "conv-delete-index-fail", "upload_id": upload_id},
            )
        finally:
            self.runtime._conversation_file_index_writer.write_index = original_write_index

        self.assertEqual(deleted.status_code, 200, deleted.text)
        marker = await self.runtime.storage.get_conversation_file_index_repair_marker("conv-delete-index-fail")
        resource = await self.runtime.storage.get_conversation_file_resource("conv-delete-index-fail", "acc-1", upload_id)
        message = await self.runtime.storage.get_message(file_upload_message_id(upload_id))
        self.assertEqual(resource.status, "deleted")
        self.assertEqual(message.metadata["file_status"], "deleted")
        self.assertEqual(marker.status, "pending")
        self.assertEqual(marker.reason_code, "delete_index_write_failed")
        self.assertEqual(marker.affected_upload_ids, (upload_id,))

    async def test_delete_index_marker_write_failure_keeps_deleted_fact_and_fails_closed(self) -> None:
        upload = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-delete-marker-fail"},
            files={"file": ("delete-me.csv", "ped_id,value\nA001,1\n", "text/csv")},
        )
        self.assertEqual(upload.status_code, 201, upload.text)
        upload_id = upload.json()["upload_id"]
        original_write_index = self.runtime._conversation_file_index_writer.write_index
        original_record_marker = self.runtime.storage.record_conversation_file_index_repair_required

        def failing_write_index(*_args, **_kwargs):
            raise OSError("index unavailable")

        async def failing_record_marker(*_args, **_kwargs):
            raise RuntimeError("marker store unavailable")

        self.runtime._conversation_file_index_writer.write_index = failing_write_index
        self.runtime.storage.record_conversation_file_index_repair_required = failing_record_marker
        try:
            deleted = await self.client.request(
                "DELETE",
                "/api/v1/conversations/uploads",
                json={"conversation_id": "conv-delete-marker-fail", "upload_id": upload_id},
            )
        finally:
            self.runtime._conversation_file_index_writer.write_index = original_write_index
            self.runtime.storage.record_conversation_file_index_repair_required = original_record_marker

        self.assertEqual(deleted.status_code, 400, deleted.text)
        resource = await self.runtime.storage.get_conversation_file_resource("conv-delete-marker-fail", "acc-1", upload_id)
        message = await self.runtime.storage.get_message(file_upload_message_id(upload_id))
        marker = await self.runtime.storage.get_conversation_file_index_repair_marker("conv-delete-marker-fail")
        self.assertEqual(resource.status, "deleted")
        self.assertEqual(message.metadata["file_status"], "deleted")
        self.assertIsNone(marker)
        audit_log = (self.workspace / "audit.jsonl").read_text(encoding="utf-8")
        self.assertIn("conversation_file.delete_index_repair_marker_failed", audit_log)

    async def test_delete_local_directory_cleanup_failure_keeps_deleted_fact(self) -> None:
        upload = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-delete-cleanup-fail"},
            files={"file": ("delete-me.csv", "ped_id,value\nA001,1\n", "text/csv")},
        )
        self.assertEqual(upload.status_code, 201, upload.text)
        upload_id = upload.json()["upload_id"]
        original_delete_resource_dir = self.runtime.conversation_file_store.delete_resource_dir

        def failing_delete_resource_dir(*_args, **_kwargs):
            raise OSError("cleanup failed")

        self.runtime.conversation_file_store.delete_resource_dir = failing_delete_resource_dir
        try:
            deleted = await self.client.request(
                "DELETE",
                "/api/v1/conversations/uploads",
                json={"conversation_id": "conv-delete-cleanup-fail", "upload_id": upload_id},
            )
        finally:
            self.runtime.conversation_file_store.delete_resource_dir = original_delete_resource_dir

        self.assertEqual(deleted.status_code, 200, deleted.text)
        resource = await self.runtime.storage.get_conversation_file_resource("conv-delete-cleanup-fail", "acc-1", upload_id)
        message = await self.runtime.storage.get_message(file_upload_message_id(upload_id))
        self.assertEqual(resource.status, "deleted")
        self.assertEqual(message.metadata["file_status"], "deleted")
        self.assertTrue((self.runtime.conversation_file_store.conversation_dir("conv-delete-cleanup-fail") / "index.md").exists())
        audit_log = (self.workspace / "audit.jsonl").read_text(encoding="utf-8")
        self.assertIn("conversation_file.upload_directory_cleanup_failed", audit_log)

    async def test_repair_marker_lazy_resolution(self) -> None:
        upload = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-lazy-repair"},
            files={"file": ("materials.csv", "ped_id,value\nA001,1\n", "text/csv")},
        )
        self.assertEqual(upload.status_code, 201, upload.text)
        await self.runtime.storage.record_conversation_file_index_repair_required(
            "conv-lazy-repair",
            reason_code="test_pending_marker",
            affected_upload_ids=(upload.json()["upload_id"],),
            now=datetime(2020, 1, 1, 0, 0, 0),
        )

        listed = await self.runtime.list_uploads("conv-lazy-repair", "acc-1")

        self.assertEqual([record.upload_id for record in listed], [upload.json()["upload_id"]])
        marker = await self.runtime.storage.get_conversation_file_index_repair_marker("conv-lazy-repair")
        self.assertEqual(marker.status, "resolved")

    async def test_description_refresh_updates_same_file_upload_message(self) -> None:
        upload = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-description-refresh"},
            files={"file": ("materials.csv", "ped_id,value\nA001,1\n", "text/csv")},
        )
        self.assertEqual(upload.status_code, 201, upload.text)
        upload_id = upload.json()["upload_id"]
        before = await self.runtime.storage.get_message(file_upload_message_id(upload_id))
        resource = await self.runtime.storage.get_conversation_file_resource("conv-description-refresh", "acc-1", upload_id)
        updated_resource = await self.runtime.storage.save_conversation_file_resource(
            replace(
                resource,
                description_status="failed",
                description_summary=None,
                updated_at=datetime(2026, 6, 16, 8, 0, 0),
            )
        )

        await self.runtime.refresh_file_upload_history_message(
            updated_resource,
            now=datetime(2026, 6, 16, 8, 0, 1),
        )

        after = await self.runtime.storage.get_message(file_upload_message_id(upload_id))
        self.assertEqual(after.message_id, before.message_id)
        self.assertEqual(after.created_at, before.created_at)
        self.assertEqual(after.updated_at, datetime(2026, 6, 16, 8, 0, 1))
        self.assertEqual(after.metadata["description_status"], "failed")
        self.assertIsNone(after.metadata["description_summary"])

    async def test_upload_txt_returns_preview_and_resolves_plain_text_for_skill(self) -> None:
        text_content = "第一行说明\n第二行说明\n"

        upload = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-txt"},
            files={"file": ("notes.txt", text_content.encode("utf-8"), "text/plain")},
        )

        self.assertEqual(upload.status_code, 201, upload.text)
        payload = upload.json()
        self.assertEqual(payload["filename"], "notes.txt")
        self.assertEqual(payload["content_type"], "text/plain")
        self.assertEqual(payload["file_type"], "text")
        self.assertEqual(payload["preview"]["shape"], "text")
        self.assertEqual(payload["preview"]["source_encoding"], "utf-8-sig")
        self.assertEqual(payload["preview"]["line_count"], 2)
        self.assertEqual(payload["preview"]["row_count"], 2)
        self.assertEqual(payload["preview"]["char_count"], len(text_content))
        self.assertEqual(payload["preview"]["normalized_content_type"], "text/plain")
        self.assertNotIn("content", payload)

        resolved = await self.runtime.resolve_uploads_for_message(
            "conv-txt", "acc-1", [payload["upload_id"]]
        )
        prompt_artifact = resolved["uploaded_artifacts"][0]
        script_artifact = resolved["skill_artifacts"][0]
        self.assertEqual(prompt_artifact["filename"], "notes.txt")
        self.assertEqual(prompt_artifact["file_type"], "text")
        self.assertNotIn("content", prompt_artifact)
        self.assertNotIn("content_base64", prompt_artifact)
        self.assertEqual(script_artifact["filename"], "notes.txt")
        self.assertEqual(script_artifact["original_filename"], "notes.txt")
        self.assertEqual(script_artifact["content_type"], "text/plain")
        self.assertEqual(script_artifact["normalized_content_type"], "text/plain")
        self.assertEqual(script_artifact["content"], text_content)
        self.assertNotIn("content_base64", script_artifact)

    async def test_submit_without_upload_ids_includes_all_active_conversation_uploads(self) -> None:
        captured_prompts: list[str] = []

        def main_agent_generator(prompt, **_kwargs):
            captured_prompts.append(str(prompt))
            return ["已收到。"]

        await self.reconfigure_runtime(main_agent_stream_generator=main_agent_generator)
        csv_upload = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-conversation-scope"},
            files={"file": ("materials.csv", "ped_id,value\nA001,1\n", "text/csv")},
        )
        text_upload = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-conversation-scope"},
            files={"file": ("notes.txt", "SECRET_NOTES\n", "text/plain")},
        )
        self.assertEqual(csv_upload.status_code, 201, csv_upload.text)
        self.assertEqual(text_upload.status_code, 201, text_upload.text)

        submitted = await self.submit_message(
            conversation_id="conv-conversation-scope",
            content="请读取当前会话里的所有文件并总结。",
            capability_id=None,
        )

        self.assertEqual(submitted.status_code, 202, submitted.text)
        task_id = submitted.json()["task_id"]
        await self.wait_for_terminal_task(task_id)
        attachments = await self.runtime.storage.list_task_input_attachments_for_task(task_id)
        self.assertEqual(attachments, [])
        self.assertTrue(captured_prompts)
        prompt = "\n".join(captured_prompts)
        self.assertIn("materials.csv", prompt)
        self.assertIn("notes.txt", prompt)
        self.assertNotIn("A001,1", prompt)
        self.assertNotIn("SECRET_NOTES", prompt)
        self.assertNotIn("content_base64", prompt)
        self.assertNotIn("storage_key", prompt)

        resolved = await self.runtime.resolve_conversation_uploads_for_message(
            "conv-conversation-scope",
            "acc-1",
        )
        self.assertEqual(
            {artifact["upload_id"] for artifact in resolved["uploaded_artifacts"]},
            {csv_upload.json()["upload_id"], text_upload.json()["upload_id"]},
        )
        self.assertEqual(len(resolved["skill_artifacts"]), 2)
        self.assertTrue(any(artifact.get("content") == "ped_id,value\nA001,1\n" for artifact in resolved["skill_artifacts"]))
        self.assertTrue(any(artifact.get("content") == "SECRET_NOTES\n" for artifact in resolved["skill_artifacts"]))

    async def test_submit_with_new_upload_id_merges_prior_conversation_uploads(self) -> None:
        captured_prompts: list[str] = []

        def main_agent_generator(prompt, **_kwargs):
            captured_prompts.append(str(prompt))
            return ["已收到。"]

        await self.reconfigure_runtime(main_agent_stream_generator=main_agent_generator)
        prior = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-merge-scope"},
            files={"file": ("prior.csv", "ped_id,value\nP001,1\n", "text/csv")},
        )
        draft = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-merge-scope"},
            files={"file": ("draft.csv", "ped_id,value\nD001,2\n", "text/csv")},
        )
        self.assertEqual(prior.status_code, 201, prior.text)
        self.assertEqual(draft.status_code, 201, draft.text)
        draft_upload_id = draft.json()["upload_id"]

        submitted = await self.submit_message(
            conversation_id="conv-merge-scope",
            content="这次用新上传的文件继续分析。",
            capability_id=None,
            metadata={"upload_ids": [draft_upload_id]},
        )

        self.assertEqual(submitted.status_code, 202, submitted.text)
        task_id = submitted.json()["task_id"]
        await self.wait_for_terminal_task(task_id)
        self.assertTrue(captured_prompts)
        prompt = "\n".join(captured_prompts)
        self.assertIn("prior.csv", prompt)
        self.assertIn("draft.csv", prompt)
        attachments = await self.runtime.storage.list_task_input_attachments_for_task(task_id)
        self.assertEqual([attachment.source_upload_id for attachment in attachments], [draft_upload_id])
        self.assertEqual(attachments[0].source_kind, "message_upload")
        self.assertNotIn("content", attachments[0].prompt_artifact)
        self.assertIn("content", attachments[0].skill_artifact)

    async def test_conversation_file_context_excludes_deleted_and_foreign_uploads(self) -> None:
        active = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-context-filter"},
            files={"file": ("active.csv", "ped_id,value\nA001,1\n", "text/csv")},
        )
        deleted = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-context-filter"},
            files={"file": ("deleted.csv", "ped_id,value\nD001,1\n", "text/csv")},
        )
        foreign = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-context-foreign"},
            files={"file": ("foreign.csv", "ped_id,value\nF001,1\n", "text/csv")},
        )
        self.assertEqual(active.status_code, 201, active.text)
        self.assertEqual(deleted.status_code, 201, deleted.text)
        self.assertEqual(foreign.status_code, 201, foreign.text)
        deleted_response = await self.client.request(
            "DELETE",
            "/api/v1/conversations/uploads",
            json={"conversation_id": "conv-context-filter", "upload_id": deleted.json()["upload_id"]},
        )
        self.assertEqual(deleted_response.status_code, 200)

        resolved = await self.runtime.resolve_conversation_uploads_for_message(
            "conv-context-filter",
            "acc-1",
        )
        self.assertEqual([artifact["filename"] for artifact in resolved["uploaded_artifacts"]], ["active.csv"])
        self.assertEqual(resolved["missing_upload_ids"], [])

        with self.assertRaises(PermissionError):
            await self.runtime.resolve_uploads_for_message(
                "conv-context-filter",
                "acc-1",
                [foreign.json()["upload_id"]],
            )

    async def test_upload_csv_normalizes_header_noise_and_preserves_original_hash(self) -> None:
        csv_bytes = "\ufeff\"ped_id\",hyb_check,set\nA001,0,A\n".encode("utf-8")

        upload = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-normalized-csv"},
            files={"file": ("materials.csv", csv_bytes, "text/csv")},
        )

        self.assertEqual(upload.status_code, 201, upload.text)
        payload = upload.json()
        self.assertEqual(payload["preview"]["columns"], ["ped_id", "hyb_check", "set"])
        self.assertEqual(payload["preview"]["source_encoding"], "utf-8-sig")
        self.assertEqual(payload["preview"]["original_columns"][0], "ped_id")
        self.assertEqual(payload["preview"]["column_normalizations"][0]["normalized"], "ped_id")
        resolved = await self.runtime.resolve_uploads_for_message(
            "conv-normalized-csv", "acc-1", [payload["upload_id"]]
        )
        self.assertEqual(
            resolved["skill_artifacts"][0]["content"].splitlines()[0],
            "ped_id,hyb_check,set",
        )
        self.assertEqual(resolved["skill_artifacts"][0]["original_filename"], "materials.csv")
        self.assertEqual(resolved["skill_artifacts"][0]["normalized_filename"], "materials.csv")
        self.assertNotIn("content", resolved["uploaded_artifacts"][0])

    async def test_upload_tsv_resolves_as_csv_family_skill_input(self) -> None:
        tsv_content = "FID\tIID\tRootAngle_deg\n0\tCML103\t46.734638\n"

        upload = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-gwas-tsv"},
            files={"file": ("phenotype.tsv", tsv_content, "text/tab-separated-values")},
        )

        self.assertEqual(upload.status_code, 201, upload.text)
        payload = upload.json()
        self.assertEqual(payload["filename"], "phenotype.tsv")
        self.assertEqual(payload["content_type"], "text/tab-separated-values")
        self.assertEqual(payload["file_type"], "csv")
        self.assertEqual(payload["preview"]["columns"], ["FID", "IID", "RootAngle_deg"])
        self.assertEqual(payload["preview"]["row_count"], 1)
        self.assertEqual(payload["preview"]["normalized_content_type"], "text/csv")
        self.assertNotIn("content", payload)

        resolved = await self.runtime.resolve_uploads_for_message(
            "conv-gwas-tsv", "acc-1", [payload["upload_id"]]
        )
        prompt_artifact = resolved["uploaded_artifacts"][0]
        script_artifact = resolved["skill_artifacts"][0]
        self.assertEqual(prompt_artifact["filename"], "phenotype.tsv")
        self.assertNotIn("content", prompt_artifact)
        self.assertNotIn("content_base64", prompt_artifact)
        self.assertEqual(script_artifact["file_type"], "csv")
        self.assertEqual(script_artifact["original_filename"], "phenotype.tsv")
        self.assertEqual(script_artifact["normalized_filename"], "phenotype.csv")
        self.assertEqual(script_artifact["filename"], "phenotype.csv")
        self.assertEqual(script_artifact["content_type"], "text/csv")
        self.assertEqual(script_artifact["normalized_content_type"], "text/csv")
        self.assertEqual(script_artifact["content"].splitlines()[0], "FID,IID,RootAngle_deg")
        self.assertIn("0,CML103,46.734638", script_artifact["content"])
        self.assertNotIn("content_base64", script_artifact)

    async def test_upload_tsv_forces_tab_delimiter_when_values_contain_commas(self) -> None:
        tsv_content = "name\tnote,raw\nCML103\talpha,beta\n"

        upload = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-tsv-commas"},
            files={"file": ("phenotype.tsv", tsv_content, "text/tab-separated-values")},
        )

        self.assertEqual(upload.status_code, 201, upload.text)
        payload = upload.json()
        self.assertEqual(payload["file_type"], "csv")
        self.assertEqual(payload["preview"]["columns"], ["name", "note,raw"])
        resolved = await self.runtime.resolve_uploads_for_message(
            "conv-tsv-commas", "acc-1", [payload["upload_id"]]
        )
        rows = list(csv.reader(StringIO(resolved["skill_artifacts"][0]["content"])))
        self.assertEqual(rows, [["name", "note,raw"], ["CML103", "alpha,beta"]])

    async def test_upload_csv_accepts_legacy_encoding_and_truncates_prompt_safe_summary(self) -> None:
        columns = [f"列{i:02d}" for i in range(55)]
        rows = [",".join(columns), ",".join(f"值{i:02d}" for i in range(55))]
        csv_bytes = ("\n".join(rows) + "\n").encode("gb18030")

        upload = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-gb18030"},
            files={"file": ("materials.csv", csv_bytes, "text/csv")},
        )

        self.assertEqual(upload.status_code, 201, upload.text)
        payload = upload.json()
        self.assertEqual(payload["preview"]["source_encoding"], "gb18030")
        self.assertEqual(payload["preview"]["column_count"], 55)
        self.assertEqual(len(payload["preview"]["columns"]), 50)
        self.assertTrue(payload["preview"]["columns_truncated"])
        resolved = await self.runtime.resolve_uploads_for_message(
            "conv-gb18030", "acc-1", [payload["upload_id"]]
        )
        self.assertIn("列54", resolved["skill_artifacts"][0]["content"])
        self.assertNotIn("值54", json.dumps(resolved["uploaded_artifacts"][0], ensure_ascii=False))

    async def test_upload_rejects_duplicate_cleaned_csv_headers(self) -> None:
        upload = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-dup"},
            files={"file": ("materials.csv", "ped_id,\ufeff\"ped_id\"\nA001,A002\n", "text/csv")},
        )

        self.assertEqual(upload.status_code, 400)
        self.assertIn("duplicate", upload.json()["detail"].lower())

    async def test_upload_json_normalizes_top_level_keys_only(self) -> None:
        payload = [{"\ufeff\"ped_id\"": "A001", "nested": {"\ufeff\"ped_id\"": "raw"}}]

        upload = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-json-normalized"},
            files={"file": ("materials.json", json.dumps(payload, ensure_ascii=False), "application/json")},
        )

        self.assertEqual(upload.status_code, 201, upload.text)
        body = upload.json()
        self.assertEqual(body["preview"]["columns"], ["ped_id", "nested"])
        resolved = await self.runtime.resolve_uploads_for_message(
            "conv-json-normalized", "acc-1", [body["upload_id"]]
        )
        normalized = json.loads(resolved["skill_artifacts"][0]["content"])
        self.assertEqual(list(normalized[0].keys()), ["ped_id", "nested"])
        self.assertIn('\ufeff"ped_id"', normalized[0]["nested"])

    async def test_upload_xlsx_single_sheet_resolves_as_normalized_csv(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Only"
        sheet.append(["\ufeff\"ped_id\"", "hyb_check", "set"])
        sheet.append(["X001", 0, "A"])
        buffer = BytesIO()
        workbook.save(buffer)

        upload = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-xlsx-single"},
            files={"file": ("materials.xlsx", buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

        self.assertEqual(upload.status_code, 201, upload.text)
        payload = upload.json()
        self.assertEqual(payload["file_type"], "spreadsheet")
        self.assertEqual(payload["preview"]["selected_sheet"], "Only")
        self.assertFalse(payload["preview"]["requires_sheet_selection"])
        resolved = await self.runtime.resolve_uploads_for_message(
            "conv-xlsx-single", "acc-1", [payload["upload_id"]]
        )
        artifact = resolved["skill_artifacts"][0]
        self.assertEqual(artifact["filename"], "materials.csv")
        self.assertEqual(artifact["original_filename"], "materials.xlsx")
        self.assertEqual(artifact["content"].splitlines()[0], "ped_id,hyb_check,set")

    async def test_upload_xls_single_sheet_resolves_as_normalized_csv(self) -> None:
        xls_bytes = base64.b64decode(_SINGLE_SHEET_XLS_BASE64)

        upload = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-xls-single"},
            files={"file": ("materials.xls", xls_bytes, "application/vnd.ms-excel")},
        )

        self.assertEqual(upload.status_code, 201, upload.text)
        payload = upload.json()
        self.assertEqual(payload["file_type"], "spreadsheet")
        self.assertEqual(payload["preview"]["selected_sheet"], "Only")
        resolved = await self.runtime.resolve_uploads_for_message(
            "conv-xls-single", "acc-1", [payload["upload_id"]]
        )
        self.assertEqual(resolved["skill_artifacts"][0]["content"].splitlines()[0], "ped_id,hyb_check,set")

    async def test_upload_xlsx_multi_sheet_requires_selection_without_executable_content(self) -> None:
        workbook = Workbook()
        workbook.active.title = "A"
        workbook.active.append(["ped_id", "hyb_check"])
        workbook.active.append(["A001", 0])
        sheet_b = workbook.create_sheet("B")
        sheet_b.append(["ped_id", "hyb_check"])
        sheet_b.append(["B001", 1])
        buffer = BytesIO()
        workbook.save(buffer)

        upload = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-xlsx-multi"},
            files={"file": ("materials.xlsx", buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

        self.assertEqual(upload.status_code, 201, upload.text)
        payload = upload.json()
        self.assertEqual(payload["file_type"], "spreadsheet")
        self.assertTrue(payload["preview"]["requires_sheet_selection"])
        self.assertEqual([item["sheet_name"] for item in payload["preview"]["excel_sheets"]], ["A", "B"])
        unresolved = await self.runtime.resolve_uploads_for_message(
            "conv-xlsx-multi", "acc-1", [payload["upload_id"]]
        )
        self.assertEqual(len(unresolved["pending_sheet_selections"]), 1)
        self.assertNotIn("content", unresolved["skill_artifacts"][0])
        self.assertNotIn("content_base64", unresolved["skill_artifacts"][0])
        stored_record = self.runtime.upload_store.get_for_message(
            upload_id=payload["upload_id"],
            username="acc-1",
            conversation_id="conv-xlsx-multi",
        )
        direct_artifact = stored_record.to_skill_artifact()
        self.assertNotIn("content", direct_artifact)
        self.assertNotIn("content_base64", direct_artifact)
        before_sheet_message = await self.runtime.storage.get_message(file_upload_message_id(payload["upload_id"]))
        selected = await self.runtime.resolve_uploads_for_message(
            "conv-xlsx-multi",
            "acc-1",
            [payload["upload_id"]],
            upload_sheet_selections={payload["upload_id"]: "B"},
        )
        self.assertEqual(selected["pending_sheet_selections"], [])
        self.assertIn("B001", selected["skill_artifacts"][0]["content"])
        after_sheet_message = await self.runtime.storage.get_message(file_upload_message_id(payload["upload_id"]))
        self.assertEqual(after_sheet_message.message_id, before_sheet_message.message_id)
        self.assertEqual(after_sheet_message.created_at, before_sheet_message.created_at)
        self.assertEqual(after_sheet_message.metadata["selected_sheet"], "B")
        self.assertNotIn("requires_sheet_selection", after_sheet_message.metadata)

    async def test_submission_upload_resolution_is_read_only_and_returns_safe_frozen_refs(self) -> None:
        workbook = Workbook()
        workbook.active.title = "Alpha"
        workbook.active.append(["ped_id"])
        workbook.active.append(["A001"])
        beta = workbook.create_sheet("Beta")
        beta.append(["ped_id"])
        beta.append(["B001"])
        buffer = BytesIO()
        workbook.save(buffer)
        upload = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-submission-upload-pure"},
            files={
                "file": (
                    "materials.xlsx",
                    buffer.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        self.assertEqual(upload.status_code, 201, upload.text)
        upload_id = upload.json()["upload_id"]

        with (
            patch.object(
                self.runtime,
                "_repair_conversation_file_index_if_due",
                new=AsyncMock(side_effect=AssertionError("index repair is not pure")),
            ) as repair,
            patch.object(
                self.runtime,
                "_apply_conversation_file_sheet_selection",
                new=AsyncMock(side_effect=AssertionError("sheet selection was persisted")),
            ) as persist_selection,
            patch.object(
                self.runtime.storage,
                "save_conversation_file_resource",
                new=AsyncMock(side_effect=AssertionError("resource was written")),
            ) as save_resource,
        ):
            resolved = await self.runtime.resolve_uploads_for_submission(
                "conv-submission-upload-pure",
                "acc-1",
                [upload_id],
                upload_sheet_selections={upload_id: "Beta"},
            )

        repair.assert_not_awaited()
        persist_selection.assert_not_awaited()
        save_resource.assert_not_awaited()
        self.assertEqual(resolved.upload_refs[0].upload_id, upload_id)
        self.assertEqual(resolved.upload_refs[0].conversation_id, "conv-submission-upload-pure")
        self.assertEqual(resolved.upload_refs[0].sha256, upload.json()["sha256"])
        self.assertEqual(resolved.upload_refs[0].size_bytes, upload.json()["size_bytes"])
        self.assertEqual(resolved.upload_refs[0].selected_sheet, "Beta")
        self.assertIn("B001", resolved.skill_artifacts[0]["content"])
        continuation = resolved.continuation_upload_refs()
        self.assertEqual(
            set(continuation[0]),
            {"upload_id", "conversation_id", "sha256", "size_bytes", "selected_sheet"},
        )
        serialized = json.dumps(continuation, ensure_ascii=False)
        for forbidden in ("content", "storage_key", "path"):
            self.assertNotIn(forbidden, serialized)
        with self.assertRaises(FrozenInstanceError):
            resolved.upload_refs[0].selected_sheet = "Alpha"
        resource = await self.runtime.storage.get_conversation_file_resource(
            "conv-submission-upload-pure", "acc-1", upload_id
        )
        self.assertIsNone(resource.selected_sheet)
        self.assertTrue(resource.requires_sheet_selection)

    async def test_submission_upload_resolution_rejects_blob_digest_drift(
        self,
    ) -> None:
        response = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-submission-blob-drift"},
            files={"file": ("materials.csv", "ped_id\nA001\n", "text/csv")},
        )
        self.assertEqual(response.status_code, 201, response.text)
        upload_id = response.json()["upload_id"]
        resource = await self.runtime.storage.get_conversation_file_resource(
            "conv-submission-blob-drift", "acc-1", upload_id
        )
        self.assertIsNotNone(resource)
        blob = self.runtime.conversation_file_store.open_path(resource.storage_key)
        blob.write_bytes(b"x" * resource.size_bytes)

        with self.assertRaisesRegex(
            UploadValidationError, "conversation_upload_blob_drift"
        ):
            await self.runtime.resolve_uploads_for_submission(
                "conv-submission-blob-drift", "acc-1", [upload_id]
            )

    async def test_submission_conversation_refs_are_deterministically_sorted(self) -> None:
        conversation_id = "conv-submission-upload-order"
        first = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": conversation_id},
            files={"file": ("first.csv", "ped_id\nA001\n", "text/csv")},
        )
        second = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": conversation_id},
            files={"file": ("second.csv", "ped_id\nB001\n", "text/csv")},
        )
        self.assertEqual(first.status_code, 201, first.text)
        self.assertEqual(second.status_code, 201, second.text)

        with patch.object(
            self.runtime,
            "_repair_conversation_file_index_if_due",
            new=AsyncMock(side_effect=AssertionError("index repair is not pure")),
        ) as repair:
            resolved = await self.runtime.resolve_conversation_uploads_for_submission(
                conversation_id,
                "acc-1",
            )

        repair.assert_not_awaited()
        continuation_ids = [item["upload_id"] for item in resolved.continuation_upload_refs()]
        self.assertEqual(continuation_ids, sorted((first.json()["upload_id"], second.json()["upload_id"])))

    async def test_submission_resolution_preserves_missing_and_cross_conversation_behavior(self) -> None:
        own = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-submission-owner"},
            files={"file": ("own.csv", "ped_id\nA001\n", "text/csv")},
        )
        self.assertEqual(own.status_code, 201, own.text)
        unknown_id = "upl-000000000000"

        legacy_missing = await self.runtime.resolve_uploads_for_message(
            "conv-submission-owner", "acc-1", [unknown_id]
        )
        pure_missing = await self.runtime.resolve_uploads_for_submission(
            "conv-submission-owner", "acc-1", [unknown_id]
        )

        self.assertEqual(pure_missing.missing_upload_ids, tuple(legacy_missing["missing_upload_ids"]))
        for resolver in (
            self.runtime.resolve_uploads_for_message,
            self.runtime.resolve_uploads_for_submission,
        ):
            with self.assertRaisesRegex(
                PermissionError,
                f"Upload does not belong to conversation: {own.json()['upload_id']}",
            ):
                await resolver("conv-submission-other", "acc-1", [own.json()["upload_id"]])

    async def test_submit_multi_sheet_upload_creates_sheet_selection_interrupt_and_resume_accepts_choice(self) -> None:
        workbook = Workbook()
        workbook.active.title = "Alpha"
        workbook.active.append(["ped_id", "hyb_check"])
        workbook.active.append(["A001", 0])
        beta = workbook.create_sheet("Beta")
        beta.append(["ped_id", "hyb_check"])
        beta.append(["B001", 1])
        buffer = BytesIO()
        workbook.save(buffer)
        upload = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-sheet-interrupt"},
            files={"file": ("materials.xlsx", buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        self.assertEqual(upload.status_code, 201, upload.text)
        upload_id = upload.json()["upload_id"]

        submitted = await self.submit_message(
            conversation_id="conv-sheet-interrupt",
            content="用这个多 sheet 文件做设计",
            capability_id=None,
            metadata={"upload_ids": [upload_id]},
        )

        self.assertEqual(submitted.status_code, 202, submitted.text)
        task_id = submitted.json()["task_id"]
        await self.wait_for_condition(
            lambda: self.runtime.storage.list_interrupts_for_task(task_id),
            timeout=3,
        )
        interrupts = await self.runtime.list_interrupts(task_id)
        open_interrupt = next(item for item in interrupts if item["status"] == "open")
        self.assertEqual(open_interrupt["reason_code"], "sheet_selection_required")
        sheet_field = open_interrupt["required_fields"]["upload_sheet_selections"]
        self.assertEqual(sheet_field["required_upload_ids"], [upload_id])
        self.assertEqual(sheet_field["options_by_upload_id"][upload_id], ["Alpha", "Beta"])
        self.assertNotIn("A001", json.dumps(sheet_field, ensure_ascii=False))
        events = await self.runtime.storage.list_events_for_task(task_id)
        waiting_events = [event for event in events if event.event_type == "node.waiting_for_input"]
        self.assertEqual(len(waiting_events), 1)
        self.assertEqual(waiting_events[0].node_id, open_interrupt["node_id"])
        self.assertEqual(waiting_events[0].payload["interrupt_id"], open_interrupt["interrupt_id"])
        self.assertEqual(waiting_events[0].payload["reason_code"], open_interrupt["reason_code"])

        invalid = await self.answer_interrupt_with_chat(
            conversation_id="conv-sheet-interrupt",
            interrupt_id=open_interrupt["interrupt_id"],
            content="选择 Missing",
            metadata={"upload_sheet_selections": {upload_id: "Missing"}},
        )
        self.assertEqual(invalid.status_code, 400)
        self.assertEqual(await self.runtime.storage.list_interrupt_answers(open_interrupt["interrupt_id"]), [])
        still_open = next(item for item in await self.runtime.list_interrupts(task_id) if item["interrupt_id"] == open_interrupt["interrupt_id"])
        self.assertEqual(still_open["status"], "open")
        node_after_invalid = await self.runtime.storage.get_task_node(open_interrupt["node_id"])
        self.assertIsNotNone(node_after_invalid)
        self.assertEqual(str(node_after_invalid.status), "waiting_for_input")

        # Re-open the scenario because the invalid answer is intentionally fail-closed.
        second = await self.submit_message(
            conversation_id="conv-sheet-interrupt-2",
            content="用这个多 sheet 文件做设计",
            capability_id=None,
            metadata={"upload_ids": [upload_id]},
        )
        self.assertEqual(second.status_code, 404, "upload is conversation scoped and must not be reusable elsewhere")

        answer = await self.answer_interrupt_with_chat(
            conversation_id="conv-sheet-interrupt",
            interrupt_id=open_interrupt["interrupt_id"],
            content="选择 Beta",
            metadata={"sheet_selections": {upload_id: "Beta"}},
        )
        self.assertEqual(answer.status_code, 202, answer.text)
        answer_payload = answer.json()
        self.assertEqual(answer_payload["action"], "interrupt_resumed")
        self.assertEqual(answer_payload["interrupt_id"], open_interrupt["interrupt_id"])
        self.assertEqual(answer_payload["answer_payload"], {"upload_sheet_selections": {upload_id: "Beta"}})
        self.assertNotIn("A001", json.dumps(answer_payload, ensure_ascii=False))
        self.assertNotIn("B001", json.dumps(answer_payload, ensure_ascii=False))
        await self.runtime._await_existing_execution(task_id)
        terminal = await self.wait_for_terminal_task(task_id)
        self.assertEqual(terminal["status"], "completed")
        saved_answers = await self.runtime.storage.list_interrupt_answers(open_interrupt["interrupt_id"])
        self.assertEqual(len(saved_answers), 1)
        self.assertEqual(saved_answers[0].answer_payload, {"upload_sheet_selections": {upload_id: "Beta"}})
        self.assertNotIn("A001", json.dumps(saved_answers[0].answer_payload, ensure_ascii=False))
        self.assertNotIn("B001", json.dumps(saved_answers[0].answer_payload, ensure_ascii=False))
        resource = await self.runtime.storage.get_conversation_file_resource("conv-sheet-interrupt", "acc-1", upload_id)
        self.assertIsNotNone(resource)
        self.assertEqual(resource.selected_sheet, "Beta")
        self.assertFalse(resource.requires_sheet_selection)

    async def test_submit_without_upload_ids_opens_sheet_selection_and_persists_conversation_sheet(self) -> None:
        workbook = Workbook()
        workbook.active.title = "Alpha"
        workbook.active.append(["ped_id", "hyb_check"])
        workbook.active.append(["A001", 0])
        beta = workbook.create_sheet("Beta")
        beta.append(["ped_id", "hyb_check"])
        beta.append(["B001", 1])
        buffer = BytesIO()
        workbook.save(buffer)
        upload = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-sheet-conversation-scope"},
            files={"file": ("materials.xlsx", buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        self.assertEqual(upload.status_code, 201, upload.text)
        upload_id = upload.json()["upload_id"]

        submitted = await self.submit_message(
            conversation_id="conv-sheet-conversation-scope",
            content="用当前会话的文件做设计",
            capability_id=None,
        )

        self.assertEqual(submitted.status_code, 202, submitted.text)
        task_id = submitted.json()["task_id"]
        await self.wait_for_condition(lambda: self.runtime.storage.list_interrupts_for_task(task_id), timeout=3)
        open_interrupt = next(item for item in await self.runtime.list_interrupts(task_id) if item["status"] == "open")
        self.assertEqual(open_interrupt["reason_code"], "sheet_selection_required")
        self.assertEqual(open_interrupt["required_fields"]["upload_sheet_selections"]["required_upload_ids"], [upload_id])

        answer = await self.answer_interrupt_with_chat(
            conversation_id="conv-sheet-conversation-scope",
            interrupt_id=open_interrupt["interrupt_id"],
            content="选择 Beta",
            metadata={"upload_sheet_selections": {upload_id: "Beta"}},
        )

        self.assertEqual(answer.status_code, 202, answer.text)
        await self.runtime._await_existing_execution(task_id)
        terminal = await self.wait_for_terminal_task(task_id)
        self.assertEqual(terminal["status"], "completed")
        resource = await self.runtime.storage.get_conversation_file_resource(
            "conv-sheet-conversation-scope",
            "acc-1",
            upload_id,
        )
        self.assertIsNotNone(resource)
        self.assertEqual(resource.selected_sheet, "Beta")
        self.assertFalse(resource.requires_sheet_selection)
        attachments = await self.runtime.storage.list_task_input_attachments_for_task(task_id)
        self.assertEqual(len(attachments), 1)
        self.assertEqual(attachments[0].source_kind, "interrupt_answer_upload")
        self.assertEqual(attachments[0].source_upload_id, upload_id)
        self.assertEqual(attachments[0].selected_sheet, "Beta")
        resolved = await self.runtime.resolve_conversation_uploads_for_message(
            "conv-sheet-conversation-scope",
            "acc-1",
        )
        self.assertEqual(resolved["pending_sheet_selections"], [])
        self.assertIn("B001", resolved["skill_artifacts"][0]["content"])
        self.assertNotIn("A001", resolved["skill_artifacts"][0]["content"])

        second = await self.submit_message(
            conversation_id="conv-sheet-conversation-scope",
            content="继续用这个文件",
            capability_id=None,
        )
        self.assertEqual(second.status_code, 202, second.text)
        second_task_id = second.json()["task_id"]
        await self.wait_for_terminal_task(second_task_id)
        self.assertEqual(await self.runtime.storage.list_interrupts_for_task(second_task_id), [])

    async def test_sheet_selection_resume_uses_task_bound_attachment_when_staged_upload_is_gone(self) -> None:
        workbook = Workbook()
        workbook.active.title = "Alpha"
        workbook.active.append(["ped_id", "hyb_check"])
        workbook.active.append(["A001", 0])
        beta = workbook.create_sheet("Beta")
        beta.append(["ped_id", "hyb_check"])
        beta.append(["B001", 1])
        buffer = BytesIO()
        workbook.save(buffer)
        upload = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-sheet-expired"},
            files={"file": ("materials.xlsx", buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        self.assertEqual(upload.status_code, 201, upload.text)
        upload_id = upload.json()["upload_id"]

        submitted = await self.submit_message(
            conversation_id="conv-sheet-expired",
            content="用这个多 sheet 文件做设计",
            capability_id=None,
            metadata={"upload_ids": [upload_id]},
        )
        self.assertEqual(submitted.status_code, 202, submitted.text)
        task_id = submitted.json()["task_id"]
        await self.wait_for_condition(
            lambda: self.runtime.storage.list_interrupts_for_task(task_id),
            timeout=3,
        )
        open_interrupt = next(item for item in await self.runtime.list_interrupts(task_id) if item["status"] == "open")
        deleted = self.runtime.upload_store.delete(
            upload_id=upload_id,
            username="acc-1",
            conversation_id="conv-sheet-expired",
        )
        self.assertTrue(deleted)

        answer = await self.answer_interrupt_with_chat(
            conversation_id="conv-sheet-expired",
            interrupt_id=open_interrupt["interrupt_id"],
            content="选择 Beta",
            metadata={"upload_sheet_selections": {upload_id: "Beta"}},
        )

        self.assertEqual(answer.status_code, 202, answer.text)
        await self.runtime._await_existing_execution(task_id)
        terminal = await self.wait_for_terminal_task(task_id)
        self.assertEqual(terminal["status"], "completed")
        interrupts_after = await self.runtime.list_interrupts(task_id)
        self.assertEqual([item["status"] for item in interrupts_after], ["answered"])
        node_after = await self.runtime.storage.get_task_node(open_interrupt["node_id"])
        self.assertIsNotNone(node_after)
        self.assertNotEqual(str(node_after.status), "waiting_for_input")
        saved_answers = await self.runtime.storage.list_interrupt_answers(open_interrupt["interrupt_id"])
        self.assertEqual(len(saved_answers), 1)
        attachments = await self.runtime.storage.list_task_input_attachments_for_task(task_id)
        self.assertEqual(len(attachments), 1)
        self.assertEqual(attachments[0].source_kind, "message_upload")
        self.assertEqual(attachments[0].selected_sheet, "Beta")
        self.assertEqual(attachments[0].interrupt_answer_id, saved_answers[0].interrupt_answer_id)
        self.assertIn("content", attachments[0].skill_artifact)
        self.assertNotIn("content", attachments[0].prompt_artifact)
        resource = await self.runtime.storage.get_conversation_file_resource("conv-sheet-expired", "acc-1", upload_id)
        self.assertIsNotNone(resource)
        self.assertEqual(resource.selected_sheet, "Beta")
        self.assertFalse(resource.requires_sheet_selection)

    async def test_vnd_ms_excel_without_excel_magic_stays_csv_compatible(self) -> None:
        upload = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-vnd-csv"},
            files={"file": ("materials", "ped_id\nA001\n", "application/vnd.ms-excel")},
        )

        self.assertEqual(upload.status_code, 201, upload.text)
        self.assertEqual(upload.json()["file_type"], "csv")

    async def test_tab_separated_mime_without_extension_stays_csv_compatible(self) -> None:
        upload = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-tsv-mime"},
            files={"file": ("phenotype", "FID\tIID\n0\tCML103\n", "text/tsv")},
        )

        self.assertEqual(upload.status_code, 201, upload.text)
        payload = upload.json()
        self.assertEqual(payload["content_type"], "text/tsv")
        self.assertEqual(payload["file_type"], "csv")
        self.assertEqual(payload["preview"]["columns"], ["FID", "IID"])
        resolved = await self.runtime.resolve_uploads_for_message(
            "conv-tsv-mime", "acc-1", [payload["upload_id"]]
        )
        self.assertEqual(resolved["skill_artifacts"][0]["normalized_filename"], "phenotype.csv")
        self.assertEqual(resolved["skill_artifacts"][0]["content"].splitlines()[0], "FID,IID")

    async def test_upload_json_returns_preview(self) -> None:
        material_data = [
            {"ped_id": "CK_A", "design_check": "1", "set": "A"},
            {"ped_id": "A001", "design_check": "0", "set": "A"},
        ]

        upload = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-json"},
            files={"file": ("materials.json", json.dumps(material_data), "application/json")},
        )

        self.assertEqual(upload.status_code, 201)
        payload = upload.json()
        self.assertEqual(payload["file_type"], "json")
        self.assertEqual(payload["preview"]["row_count"], 2)
        self.assertEqual(payload["preview"]["columns"], ["ped_id", "design_check", "set"])

    async def test_upload_png_resolves_binary_content_only_for_skill_scripts(self) -> None:
        png_content = b"\x89PNG\r\n\x1a\nocr-test"

        upload = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-ocr"},
            files={"file": ("scan.png", png_content, "image/png")},
        )

        self.assertEqual(upload.status_code, 201)
        payload = upload.json()
        self.assertEqual(payload["filename"], "scan.png")
        self.assertEqual(payload["file_type"], "image")
        self.assertEqual(payload["preview"]["shape"], "binary")
        self.assertEqual(payload["preview"]["columns"], [])
        self.assertIsNone(payload["preview"]["row_count"])
        self.assertNotIn("content", payload)
        self.assertNotIn("content_base64", payload)

        resolved = await self.runtime.resolve_uploads_for_message(
            "conv-ocr", "acc-1", [payload["upload_id"]]
        )
        prompt_artifact = resolved["uploaded_artifacts"][0]
        script_artifact = resolved["skill_artifacts"][0]
        self.assertEqual(prompt_artifact["filename"], "scan.png")
        self.assertNotIn("content", prompt_artifact)
        self.assertNotIn("content_base64", prompt_artifact)
        self.assertEqual(script_artifact["content_base64"], base64.b64encode(png_content).decode("ascii"))
        self.assertEqual(script_artifact["encoding"], "base64")
        self.assertNotIn("content", script_artifact)

    async def test_upload_vcf_resolves_binary_content_only_for_skill_scripts(self) -> None:
        vcf_content = (
            b"##fileformat=VCFv4.2\n"
            b"#CHROM\tPOS\tID\tREF\tALT\tQUAL\tFILTER\tINFO\tFORMAT\tsample_1\n"
            b"1\t42\t.\tA\tG\t.\tPASS\t.\tGT\t0/1\n"
        )
        self.runtime.upload_store.max_preview_bytes = 8

        upload = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-rice-vcf"},
            files={"file": ("sample.vcf", vcf_content, "text/plain")},
        )

        self.assertEqual(upload.status_code, 201, upload.text)
        payload = upload.json()
        self.assertEqual(payload["filename"], "sample.vcf")
        self.assertEqual(payload["file_type"], "vcf")
        self.assertEqual(payload["preview"]["shape"], "binary")
        self.assertIsNone(payload["preview"]["row_count"])
        self.assertNotIn("content", payload)
        self.assertNotIn("content_base64", payload)

        resolved = await self.runtime.resolve_uploads_for_message(
            "conv-rice-vcf", "acc-1", [payload["upload_id"]]
        )
        prompt_artifact = resolved["uploaded_artifacts"][0]
        script_artifact = resolved["skill_artifacts"][0]
        self.assertEqual(prompt_artifact["filename"], "sample.vcf")
        self.assertEqual(prompt_artifact["file_type"], "vcf")
        self.assertNotIn("content", prompt_artifact)
        self.assertNotIn("content_base64", prompt_artifact)
        self.assertEqual(script_artifact["filename"], "sample.vcf")
        self.assertEqual(script_artifact["content_base64"], base64.b64encode(vcf_content).decode("ascii"))
        self.assertEqual(script_artifact["encoding"], "base64")
        self.assertNotIn("content", script_artifact)

    async def test_upload_vcf_gz_resolves_by_compound_extension_without_accepting_plain_gz(self) -> None:
        compressed_vcf = gzip.compress(
            b"##fileformat=VCFv4.2\n"
            b"#CHROM\tPOS\tID\tREF\tALT\tQUAL\tFILTER\tINFO\tFORMAT\tsample_1\n"
            b"1\t42\t.\tA\tG\t.\tPASS\t.\tGT\t0/1\n"
        )

        upload = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-rice-vcf-gz"},
            files={"file": ("sample.vcf.gz", compressed_vcf, "application/gzip")},
        )

        self.assertEqual(upload.status_code, 201, upload.text)
        payload = upload.json()
        self.assertEqual(payload["filename"], "sample.vcf.gz")
        self.assertEqual(payload["file_type"], "vcf")
        resolved = await self.runtime.resolve_uploads_for_message(
            "conv-rice-vcf-gz", "acc-1", [payload["upload_id"]]
        )
        self.assertEqual(resolved["skill_artifacts"][0]["filename"], "sample.vcf.gz")
        self.assertEqual(
            resolved["skill_artifacts"][0]["content_base64"],
            base64.b64encode(compressed_vcf).decode("ascii"),
        )

        plain_gz = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-rice-vcf-gz"},
            files={"file": ("archive.gz", compressed_vcf, "application/gzip")},
        )
        self.assertEqual(plain_gz.status_code, 400)


    async def test_image_upload_is_persistent_without_description(self) -> None:
        png_bytes = b"\x89PNG\r\n\x1a\n" + b"0" * 16
        upload = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-image-index"},
            files={"file": ("leaf.png", png_bytes, "image/png")},
        )

        self.assertEqual(upload.status_code, 201, upload.text)
        upload_id = upload.json()["upload_id"]
        resource = await self.runtime.storage.get_conversation_file_resource("conv-image-index", "acc-1", upload_id)
        self.assertEqual(resource.description_status, "not_required")
        index_text = (self.runtime.conversation_file_store.root_dir / "conv-image-index" / "index.md").read_text(encoding="utf-8")
        self.assertIn("图片文件不自动生成描述", index_text)

    async def test_upload_rejects_conversation_ids_that_cannot_be_safely_stored(self) -> None:
        conversation_id = "conv-" + ("x" * 300)

        upload = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": conversation_id},
            files={"file": ("materials.csv", "ped_id\nA\n", "text/csv")},
        )

        self.assertEqual(upload.status_code, 400)
        self.assertIn("conversation_id failed file storage safety validation", upload.json()["detail"])
        self.assertIsNone(await self.runtime.storage.get_conversation(conversation_id))

    async def test_list_and_delete_uploads_for_conversation(self) -> None:
        first = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-files"},
            files={"file": ("first.csv", "ped_id,design_check\nA,0\n", "text/csv")},
        )
        second = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-files"},
            files={"file": ("second.json", json.dumps([{"ped_id": "B", "design_check": "0"}]), "application/json")},
        )
        self.assertEqual(first.status_code, 201)
        self.assertEqual(second.status_code, 201)

        listed = await self.client.get("/api/v1/conversations/conv-files/uploads")
        self.assertEqual(listed.status_code, 200)
        payload = listed.json()
        self.assertEqual(payload["conversation_id"], "conv-files")
        self.assertEqual([item["filename"] for item in payload["uploads"]], ["first.csv", "second.json"])
        self.assertNotIn("content", payload["uploads"][0])
        first_upload_id = first.json()["upload_id"]
        first_resource_dir = self.runtime.conversation_file_store.root_dir / "conv-files" / first_upload_id
        stored_original = first_resource_dir / "original"
        stored_description = first_resource_dir / "description.json"
        index_path = self.runtime.conversation_file_store.root_dir / "conv-files" / "index.md"
        self.assertEqual(stored_original.read_text(encoding="utf-8"), "ped_id,design_check\nA,0\n")
        self.assertTrue(stored_description.exists())
        self.assertIn(f"{first_upload_id} — first.csv", index_path.read_text(encoding="utf-8"))
        stored_resource = await self.runtime.storage.get_conversation_file_resource("conv-files", "acc-1", first_upload_id)
        self.assertIsNotNone(stored_resource)
        self.assertEqual(stored_resource.description_status, "ready")

        paged = await self.client.get("/api/v1/conversations/conv-files/uploads?limit=1")
        self.assertEqual(len(paged.json()["uploads"]), 1)
        self.assertEqual(paged.json()["next_cursor"], first_upload_id)

        deleted = await self.client.request(
            "DELETE",
            "/api/v1/conversations/uploads",
            json={"conversation_id": "conv-files", "upload_id": first.json()["upload_id"]},
        )
        self.assertEqual(deleted.status_code, 200)
        self.assertEqual(deleted.json()["deleted"], True)
        self.assertFalse(first_resource_dir.exists())

        listed_after_delete = await self.client.get("/api/v1/conversations/conv-files/uploads")
        self.assertEqual([item["filename"] for item in listed_after_delete.json()["uploads"]], ["second.json"])
        listed_with_deleted = await self.client.get("/api/v1/conversations/conv-files/uploads?include_deleted=true")
        deleted_items = [item for item in listed_with_deleted.json()["uploads"] if item["upload_id"] == first_upload_id]
        self.assertEqual(deleted_items[0]["status"], "deleted")
        self.assertIn("文件本体已物理删除", index_path.read_text(encoding="utf-8"))

        resolved = await self.runtime.resolve_uploads_for_message(
            "conv-files", "acc-1", [first.json()["upload_id"]]
        )
        self.assertEqual(resolved["uploaded_artifacts"], [])
        self.assertEqual(resolved["missing_upload_ids"], [first.json()["upload_id"]])

    async def test_submit_with_missing_upload_id_fails_closed_before_creating_task(self) -> None:
        submitted = await self.submit_message(
            conversation_id="conv-missing-upload-submit",
            content="用这个不存在的文件做设计",
            capability_id=None,
            metadata={"upload_ids": ["upl-missing"]},
        )

        self.assertEqual(submitted.status_code, 400)
        self.assertIn("Missing or expired uploads", submitted.json()["detail"])
        self.assertIsNone(await self.runtime.storage.get_conversation("conv-missing-upload-submit"))

    async def test_missing_conversation_upload_list_and_unknown_delete_are_documented_noops(self) -> None:
        listed = await self.client.get("/api/v1/conversations/missing-conversation/uploads")

        self.assertEqual(listed.status_code, 200)
        self.assertEqual(
            listed.json(),
            {"conversation_id": "missing-conversation", "uploads": []},
        )

        deleted = await self.client.request(
            "DELETE",
            "/api/v1/conversations/uploads",
            json={"conversation_id": "missing-conversation", "upload_id": "upl-missing"},
        )

        self.assertEqual(deleted.status_code, 200)
        self.assertEqual(deleted.json(), {"upload_id": "upl-missing", "deleted": False})

    async def test_upload_rejects_unsupported_file_type(self) -> None:
        response = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-upload"},
            files={"file": ("notes.md", "hello", "text/markdown")},
        )

        self.assertEqual(response.status_code, 400)

    async def test_upload_rejects_extra_identity_and_auth_form_fields(self) -> None:
        for field_name in ("username", "auth_token", "session", "captcha", "password", "unexpected"):
            with self.subTest(field_name=field_name):
                response = await self.client.post(
                    "/api/v1/conversations/uploads",
                    data={"conversation_id": "conv-upload", field_name: "spoof"},
                    files={"file": ("materials.csv", "ped_id,design_check\nA,0\n", "text/csv")},
                )

                self.assertEqual(response.status_code, 422, response.text)

    async def test_upload_rejects_oversized_file_with_configured_limit(self) -> None:
        self.runtime.upload_store.max_file_bytes = 16

        response = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-upload"},
            files={"file": ("large.csv", "col\n" + "x" * 32, "text/csv")},
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("exceeds 16 bytes", response.json()["detail"])

    async def test_upload_and_reference_are_owner_scoped(self) -> None:
        upload = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-owned"},
            files={"file": ("data.csv", "ped_id,design_check\nA,0\n", "text/csv")},
        )
        self.assertEqual(upload.status_code, 201)
        upload_id = upload.json()["upload_id"]

        await self.login("bob")

        forbidden_upload = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-owned"},
            files={"file": ("data.csv", "ped_id,design_check\nB,0\n", "text/csv")},
        )
        self.assertEqual(forbidden_upload.status_code, 404)

        self.runtime.upload_store.max_file_bytes = 16
        forbidden_oversized_upload = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": "conv-owned"},
            files={"file": ("large.csv", "col\n" + "x" * 32, "text/csv")},
        )
        self.assertEqual(forbidden_oversized_upload.status_code, 404)

        submitted = await self.submit_message(
            conversation_id="conv-owned",
            content="用文件做RCBD",
            capability_id=None,
            metadata={"upload_ids": [upload_id], "blocks": 3},
        )
        self.assertEqual(submitted.status_code, 404)


if __name__ == "__main__":
    unittest.main()


_SINGLE_SHEET_XLS_BASE64 = (
    '0M8R4KGxGuEAAAAAAAAAAAAAAAAAAAAAPgADAP7/CQAGAAAAAAAAAAAAAAABAAAACQAAAAAAAAAAEAAA/v///wAAAAD+////'
    'AAAAAAgAAAD/////////////////////////////////////////////////////////////////////////////////////'
    '////////////////////////////////////////////////////////////////////////////////////////////////'
    '////////////////////////////////////////////////////////////////////////////////////////////////'
    '////////////////////////////////////////////////////////////////////////////////////////////////'
    '////////////////////////////////////////////////////////////////////////////////////////////////'
    '////////////////////////////////////////////////////////////////////////////////////////////////'
    '//////////8JCBAAAAYFALsNzAcAAAAABgAAAOEAAgCwBMEAAgAAAOIAAABcAHAATm9uZSAgICAgICAgICAgICAgICAgICAg'
    'ICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAg'
    'ICAgICAgICAgICAgICAgIEIAAgCwBGEBAgAAAD0BAgABAJwAAgAOABkAAgAAABIAAgAAAGMAAgAAABMAAgAAAK8BAgAAALwB'
    'AgAAAEAAAgAAAI0AAgAAAD0AEgDgAVoAzz9OKjgAAAAAAAEAWAIiAAIAAAAOAAIAAQC3AQIAAADaAAIAAAAxABUAyAAAAP9/'
    'kAEAAAAAAQAFAEFyaWFsMQAVAMgAAAD/f5ABAAAAAAEABQBBcmlhbDEAFQDIAAAA/3+QAQAAAAABAAUAQXJpYWwxABUAyAAA'
    'AP9/kAEAAAAAAQAFAEFyaWFsMQAVAMgAAAD/f5ABAAAAAAEABQBBcmlhbDEAFQDIAAAA/3+QAQAAAAABAAUAQXJpYWwxABUA'
    'yAAAAP9/kAEAAAAAAQAFAEFyaWFsHgQMAKQABwAAR2VuZXJhbOAAFAAGAKQA9f8gAAD0AAAAAAAAAADAIOAAFAAGAKQA9f8g'
    'AAD0AAAAAAAAAADAIOAAFAAGAKQA9f8gAAD0AAAAAAAAAADAIOAAFAAGAKQA9f8gAAD0AAAAAAAAAADAIOAAFAAGAKQA9f8g'
    'AAD0AAAAAAAAAADAIOAAFAAGAKQA9f8gAAD0AAAAAAAAAADAIOAAFAAGAKQA9f8gAAD0AAAAAAAAAADAIOAAFAAGAKQA9f8g'
    'AAD0AAAAAAAAAADAIOAAFAAGAKQA9f8gAAD0AAAAAAAAAADAIOAAFAAGAKQA9f8gAAD0AAAAAAAAAADAIOAAFAAGAKQA9f8g'
    'AAD0AAAAAAAAAADAIOAAFAAGAKQA9f8gAAD0AAAAAAAAAADAIOAAFAAGAKQA9f8gAAD0AAAAAAAAAADAIOAAFAAGAKQA9f8g'
    'AAD0AAAAAAAAAADAIOAAFAAGAKQA9f8gAAD0AAAAAAAAAADAIOAAFAAGAKQA9f8gAAD0AAAAAAAAAADAIOAAFAAGAKQAAQAg'
    'AAD4AAAAAAAAAADAIOAAFAAHAKQAAQAgAAD4AAAAAAAAAADAIJMCBAAAgAD/YAECAAEAhQAMANEDAAAAAAQAT25sefwALgAF'
    'AAAABQAAAAYAAHBlZF9pZAkAAGh5Yl9jaGVjawMAAHNldAQAAFgwMDEBAABBCgAAAAkIEAAABhAAuw3MBwAAAAAGAAAADQAC'
    'AAEADAACAGQADwACAAEAEQACAAAAEAAIAPyp8dJNYlA/XwACAAAAgAAIAAAAAAABAAAAJQIEAAAA/wCBAAIAAQwAAg4AAAAA'
    'AAIAAAAAAAMAAAAqAAIAAAArAAIAAACCAAIAAQAbAAIAAAAaAAIAAAAUAAUAAgAAJlAVAAUAAgAAJkaDAAIAAQCEAAIAAAAm'
    'AAgAMzMzMzMz0z8nAAgAMzMzMzMz0z8oAAgAhetRuB6F4z8pAAgArkfhehSu1z+hACIACQBkAAEAAQABAIMALAEsAZqZmZmZ'
    'mbk/mpmZmZmZuT8BABIAAgAAAN0AAgAAABkAAgAAAGMAAgAAABMAAgAAAAgCEAAAAAAAAwD/AAAAAAAAAQ8A/QAKAAAAAAAR'
    'AAAAAAD9AAoAAAABABEAAQAAAP0ACgAAAAIAEQACAAAACAIQAAEAAAADAP8AAAAAAAABDwD9AAoAAQAAABEAAwAAAH4CCgAB'
    'AAEAEQACAAAA/QAKAAEAAgARAAQAAAA+AhIAtgIAAAAAQAAAAAAAAAAAAAAACgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AQAAAAIAAAADAAAABAAAAAUAAAAGAAAABwAAAP7////9/////v//////////////////////////////////////////////'
    '////////////////////////////////////////////////////////////////////////////////////////////////'
    '////////////////////////////////////////////////////////////////////////////////////////////////'
    '////////////////////////////////////////////////////////////////////////////////////////////////'
    '////////////////////////////////////////////////////////////////////////////////////////////////'
    '////////////////////////////////////////////////////////////////////////////////////////////////'
    '////////////////////////////////////////////////////////////////////////////////////////////////'
    '//////////9SAG8AbwB0ACAARQBuAHQAcgB5AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'FgAFAf//////////AQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAP7///8AAAAAAAAAAFcAbwByAGsA'
    'YgBvAG8AawAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAASAAIB////////////////'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAH///////////////8AAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAD+////AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAf///////////////wAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAP7///8AAAAAAAAAAA=='
)
