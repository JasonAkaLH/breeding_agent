from __future__ import annotations

import json
import textwrap
from datetime import datetime, timedelta
from io import BytesIO
from unittest.mock import AsyncMock, patch

from openpyxl import Workbook

from src.api.dto import SubmitMessageRequest
from src.api.file_selection import (
    FileSelectionDecision,
    FileRequirementProfile,
    FileRequirementProfileError,
    FileSelectionAnswerResolver,
    FileSelectionTriggerDetector,
    build_recent_usage,
    candidate_from_resource,
    deterministic_file_decision,
)
from src.api.file_selection_runtime import ConversationFileSelectionRuntimeMixin
from src.api.runtime import ApiRuntime
from src.core.enums import EventVisibility, MessageRole
from src.core.models import Task, TaskInputAttachment

from tests.api.support import APITestCase


class ConversationFileSelectionAPITest(APITestCase):
    def test_file_selection_keeps_one_api_domain_and_runtime_owner(self) -> None:
        self.assertIn(ConversationFileSelectionRuntimeMixin, ApiRuntime.__mro__)
        self.assertEqual(
            ConversationFileSelectionRuntimeMixin.__module__,
            "src.api.file_selection_runtime",
        )
        self.assertEqual(FileSelectionDecision.__module__, "src.api.file_selection")

    async def asyncSetUp(self) -> None:
        await super().asyncSetUp()
        await self.reconfigure_runtime(main_agent_stream_generator=lambda _prompt: ["完成。"])


    def _write_file_required_skill(self, root_name: str = "file-required-skill") -> str:
        skill_dir = self.workspace / root_name / "file-required"
        (skill_dir / "schemas").mkdir(parents=True)
        (skill_dir / "scripts").mkdir()
        (skill_dir / "SKILL.md").write_text(
            textwrap.dedent(
                """\
                ---
                name: file-required
                description: File required fixture.
                ---
                # File Required
                """
            ),
            encoding="utf-8",
        )
        (skill_dir / "skill.contract.yaml").write_text(
            textwrap.dedent(
                """\
                contract_version: '2'
                capability:
                  id: skill.file_required
                  display_name: File Required
                  description: Requires a file.
                routing:
                  triggers: [file required]
                file_selection:
                  required: true
                  expected_content: [材料文件]
                  supported_file_types: [csv, txt]
                runtime:
                  mode: python_subprocess
                entrypoints:
                  run:
                    path: scripts/run.py
                    input_schema: file_input
                    output: file_output
                input_schemas:
                  file_input:
                    path: schemas/file.input.yaml
                    title: File input
                outputs:
                  file_output:
                    required: [response_text]
                """
            ),
            encoding="utf-8",
        )
        (skill_dir / "schemas" / "file.input.yaml").write_text(
            textwrap.dedent(
                """\
                schema_id: file_input
                inputs:
                  material_file:
                    type: artifact
                    required: true
                    description: 材料文件
                    file_selection:
                      required: true
                      expected_content: [材料表]
                      supported_file_types: [csv]
                      helpful_columns: [ped_id]
                      disambiguation_hint: 优先选择材料表。
                """
            ),
            encoding="utf-8",
        )
        (skill_dir / "scripts" / "run.py").write_text("print('ok')\n", encoding="utf-8")
        return "skill.file_required"

    async def _upload_csv(self, conversation_id: str, filename: str, content: str = "ped_id,value\nA001,1\n") -> str:
        response = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": conversation_id},
            files={"file": (filename, content, "text/csv")},
        )
        self.assertEqual(response.status_code, 201, response.text)
        return response.json()["upload_id"]

    def _assert_audit_payload_redacted(self, payload: object) -> None:
        serialized = json.dumps(payload, ensure_ascii=False, default=str)
        for forbidden in (
            "storage_key",
            "content_base64",
            "mount_path",
            "provider_payload",
            "raw_payload",
            "raw_prompt",
            "raw_output",
            "user_message",
            "SECRET_VALUE",
            "token=abc",
            "api_key",
            "Authorization",
            "Bearer abc",
            "/tmp/",
            "/Users/",
        ):
            self.assertNotIn(forbidden, serialized)

    async def _upload_xlsx(
        self,
        conversation_id: str,
        filename: str,
        rows: list[list[object]] | None = None,
    ) -> str:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "materials"
        for row in rows or [["材料名称", "是否对照", "组别"], ["A001", 0, 1]]:
            sheet.append(row)
        return await self._upload_workbook(conversation_id, filename, workbook)

    async def _upload_multi_sheet_xlsx(self, conversation_id: str, filename: str) -> str:
        workbook = Workbook()
        workbook.active.title = "Alpha"
        workbook.active.append(["ped_id", "hyb_check"])
        workbook.active.append(["A001", 0])
        beta = workbook.create_sheet("Beta")
        beta.append(["ped_id", "hyb_check"])
        beta.append(["B001", 1])
        return await self._upload_workbook(conversation_id, filename, workbook)

    async def _upload_workbook(self, conversation_id: str, filename: str, workbook: Workbook) -> str:
        buffer = BytesIO()
        workbook.save(buffer)
        response = await self.client.post(
            "/api/v1/conversations/uploads",
            data={"conversation_id": conversation_id},
            files={
                "file": (
                    filename,
                    buffer.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        self.assertEqual(response.status_code, 201, response.text)
        return response.json()["upload_id"]

    async def test_active_conversation_file_is_available_without_task_attachment(self) -> None:
        self.runtime._conversation_file_selector_mode = "enforce_narrow"
        conversation_id = "conv-file-auto"
        upload_id = await self._upload_csv(conversation_id, "materials.csv")

        response = await self.submit_message(
            conversation_id=conversation_id,
            capability_id=None,
            content="请用刚才上传的文件做一个摘要。",
        )

        self.assertEqual(response.status_code, 202, response.text)
        task_id = response.json()["task_id"]
        await self.runtime._await_existing_execution(task_id)
        attachments = await self.runtime.storage.list_task_input_attachments_for_task(task_id)
        self.assertEqual(attachments, [])
        resolved = await self.runtime.resolve_conversation_uploads_for_message(conversation_id, "acc-1")
        self.assertEqual([artifact["upload_id"] for artifact in resolved["uploaded_artifacts"]], [upload_id])
        events = await self.runtime.storage.list_events_for_task(task_id)
        self.assertNotIn("conversation_file.file_selector_auto_bound", [event.event_type for event in events])

    async def test_referenced_original_spreadsheet_filename_binds_file_selector_attachment(self) -> None:
        self.runtime._conversation_file_selector_mode = "enforce_narrow"
        conversation_id = "conv-file-original-name"
        selected_id = await self._upload_xlsx(conversation_id, "间比法双组材料清单.xlsx")
        await self._upload_xlsx(conversation_id, "对角线增广列数20材料清单.xlsx")

        response = await self.submit_message(
            conversation_id=conversation_id,
            capability_id=None,
            content="那你用“间比法双组材料清单.xlsx”给我做个间比法试验设计。",
        )

        self.assertEqual(response.status_code, 202, response.text)
        task_id = response.json()["task_id"]
        await self.runtime._await_existing_execution(task_id)
        attachments = await self.runtime.storage.list_task_input_attachments_for_task(task_id)
        self.assertEqual(len(attachments), 1)
        self.assertEqual(attachments[0].source_kind, "file_selector")
        self.assertEqual(attachments[0].source_upload_id, selected_id)
        events = await self.runtime.storage.list_events_for_task(task_id)
        auto_bound = next(event for event in events if event.event_type == "conversation_file.file_selector_auto_bound")
        self.assertEqual(auto_bound.payload["selected_upload_ids"], [selected_id])

    async def test_prepared_agent_request_exposes_only_selector_winner_artifacts(
        self,
    ) -> None:
        self.runtime._conversation_file_selector_mode = "enforce_narrow"
        conversation_id = "conv-prepared-selector-winner-only"
        selected_id = await self._upload_csv(
            conversation_id, "selected.csv", "ped_id,value\nA001,1\n"
        )
        await self._upload_csv(
            conversation_id, "unselected.csv", "ped_id,value\nB001,2\n"
        )
        captured_requests = []
        initialize_run = self.runtime.agent_loop_orchestrator.initialize_run

        async def capture_request(request):
            captured_requests.append(request)
            return await initialize_run(request)

        self.runtime.agent_loop_orchestrator.initialize_run = capture_request

        response = await self.submit_message(
            conversation_id=conversation_id,
            capability_id=None,
            content="请分析 selected.csv。",
        )

        self.assertEqual(response.status_code, 202, response.text)
        self.assertEqual(len(captured_requests), 1)
        metadata = captured_requests[0].metadata
        self.assertEqual(metadata["upload_ids"], [selected_id])
        self.assertEqual(
            [item["upload_id"] for item in metadata["uploaded_artifacts"]],
            [selected_id],
        )
        self.assertEqual(
            [item["upload_id"] for item in metadata["skill_artifacts"]],
            [selected_id],
        )

    async def test_file_interrupt_exact_replay_is_independent_of_selector_cache(
        self,
    ) -> None:
        capability_id = self._write_file_required_skill()
        skill_root = self.workspace / "file-required-skill"
        await self.reconfigure_runtime(
            skill_roots=(skill_root,), public_skill_roots=(skill_root,)
        )
        self.runtime._conversation_file_selector_mode = "enforce_narrow"
        captured = []
        materialize = self.runtime.materialize_interrupt_handoff

        async def capture_materialization(record, prepared):
            captured.append((record, prepared))
            return await materialize(record, prepared)

        self.runtime.materialize_interrupt_handoff = capture_materialization
        response = await self.submit_message(
            conversation_id="conv-file-interrupt-exact-replay",
            capability_id=capability_id,
            content="执行这个 Skill。",
        )
        self.assertEqual(response.status_code, 202, response.text)
        self.assertEqual(len(captured), 1)

        self.runtime._submission_selector_facts.clear()
        self.runtime._submission_file_selection_computations.clear()

        replayed = await materialize(*captured[0])

        self.assertEqual(replayed.kind, "interrupt")

    async def test_initial_message_ordinal_reference_binds_file_selector_attachment(self) -> None:
        self.runtime._conversation_file_selector_mode = "enforce_narrow"
        conversation_id = "conv-file-ordinal"
        selected_id = await self._upload_xlsx(conversation_id, "对角线增广列数20材料清单.xlsx")
        await self._upload_xlsx(conversation_id, "间比法双组材料清单.xlsx")

        response = await self.submit_message(
            conversation_id=conversation_id,
            capability_id=None,
            content="那你用第一份文件帮我做一个对角线增广试验吧。",
        )

        self.assertEqual(response.status_code, 202, response.text)
        task_id = response.json()["task_id"]
        await self.runtime._await_existing_execution(task_id)
        attachments = await self.runtime.storage.list_task_input_attachments_for_task(task_id)
        self.assertEqual(len(attachments), 1)
        self.assertEqual(attachments[0].source_kind, "file_selector")
        self.assertEqual(attachments[0].source_upload_id, selected_id)

    async def test_shadow_records_audit_but_does_not_bind_or_interrupt(self) -> None:
        self.runtime._conversation_file_selector_mode = "shadow"
        conversation_id = "conv-file-shadow"
        await self._upload_csv(conversation_id, "materials.csv")

        response = await self.submit_message(
            conversation_id=conversation_id,
            capability_id=None,
            content="请用刚才上传的文件做摘要。",
        )

        self.assertEqual(response.status_code, 202, response.text)
        task_id = response.json()["task_id"]
        await self.runtime._await_existing_execution(task_id)
        self.assertEqual(await self.runtime.storage.list_interrupts_for_task(task_id), [])
        self.assertEqual(await self.runtime.storage.list_task_input_attachments_for_task(task_id), [])
        events = await self.runtime.storage.list_events_for_task(task_id)
        event_types = [event.event_type for event in events]
        self.assertIn("conversation_file.file_selector_invoked", event_types)
        self.assertIn("conversation_file.file_selector_decision_recorded", event_types)
        decision = next(event for event in events if event.event_type == "conversation_file.file_selector_decision_recorded")
        self.assertFalse(decision.payload["would_clarify"])
        self.assertTrue(decision.payload["would_auto_bind"])
        self.assertNotIn("user_message", json.dumps(decision.payload, ensure_ascii=False))
        self.assertNotIn("content_base64", json.dumps(decision.payload, ensure_ascii=False))
        self.assertNotIn("storage_key", json.dumps(decision.payload, ensure_ascii=False))

    async def test_file_selector_compute_is_pure_and_matches_legacy_decision(self) -> None:
        selector_calls = 0

        def selector_generator(_prompt: str, **_kwargs) -> str:
            nonlocal selector_calls
            selector_calls += 1
            return json.dumps(
                {
                    "decision": "select_one",
                    "upload_ids": [selected_id],
                    "confidence": 0.93,
                    "reason_code": "llm_selected",
                }
            )

        await self.reconfigure_runtime(skill_input_text_generator=selector_generator)
        self.runtime._conversation_file_selector_mode = "enforce_narrow"
        conversation_id = "conv-file-selector-compute-pure"
        selected_id = await self._upload_csv(conversation_id, "materials.csv")
        await self._upload_csv(conversation_id, "other.csv")
        task = Task("task-file-selector-compute-pure", conversation_id, "msg-file-selector-compute-pure")
        request = SubmitMessageRequest(
            conversation_id=conversation_id,
            content="请分析材料文件。",
            metadata={"file_requirement_profile": {"required": True}},
        )
        metadata = dict(request.metadata)

        with (
            patch.object(
                self.runtime,
                "_record_file_selection_audit_event",
                new=AsyncMock(side_effect=AssertionError("audit was persisted")),
            ) as audit,
            patch.object(
                self.runtime,
                "_bind_file_selection_uploads_or_open_sheet_selection",
                new=AsyncMock(side_effect=AssertionError("attachment was bound")),
            ) as bind,
            patch.object(
                self.runtime,
                "_open_file_selection_interrupt",
                new=AsyncMock(side_effect=AssertionError("interrupt was opened")),
            ) as open_interrupt,
            patch.object(
                self.runtime,
                "_open_sheet_selection_interrupt",
                new=AsyncMock(side_effect=AssertionError("sheet interrupt was opened")),
            ) as open_sheet_interrupt,
            patch.object(
                self.runtime.interrupt_service,
                "open_interrupt",
                new=AsyncMock(side_effect=AssertionError("interrupt service was written")),
            ) as interrupt_service,
            patch.object(
                self.runtime,
                "_repair_conversation_file_index_if_due",
                new=AsyncMock(side_effect=AssertionError("index repair was attempted")),
            ) as repair,
            patch.object(
                self.runtime,
                "_apply_conversation_file_sheet_selection",
                new=AsyncMock(side_effect=AssertionError("sheet selection was persisted")),
            ) as persist_selection,
        ):
            computed = await self.runtime._compute_conversation_file_selection(
                task=task,
                username="acc-1",
                request=request,
                metadata=metadata,
            )

        self.assertIsNotNone(computed)
        self.assertTrue(computed.triggered)
        self.assertEqual(computed.decision.upload_ids, (selected_id,))
        self.assertEqual(computed.decision.reason_code, "llm_selected")
        legacy_decision = await self.runtime._conversation_file_selection_decision(
            text=request.content,
            profile=computed.profile,
            candidates=computed.candidates,
            metadata=metadata,
        )
        self.assertEqual(computed.decision, legacy_decision)
        self.assertEqual(selector_calls, 2)
        self.assertEqual(
            computed.invoked_payload,
            self.runtime._file_selection_invoked_payload(
                mode=computed.mode,
                trigger_reason=computed.trigger_reason,
                profile=computed.profile,
                candidates=computed.candidates,
            ),
        )
        self.assertEqual(
            computed.decision_payload,
            self.runtime._file_selection_decision_payload(
                mode=computed.mode,
                trigger_reason=computed.trigger_reason,
                profile=computed.profile,
                candidates=computed.candidates,
                decision=legacy_decision,
            ),
        )
        audit.assert_not_awaited()
        bind.assert_not_awaited()
        open_interrupt.assert_not_awaited()
        open_sheet_interrupt.assert_not_awaited()
        interrupt_service.assert_not_awaited()
        repair.assert_not_awaited()
        persist_selection.assert_not_awaited()

    async def test_selector_audit_release_catalog_is_audit_only_and_redacted(self) -> None:
        auto_conversation = "conv-file-audit-catalog-auto"
        auto_upload = await self._upload_csv(auto_conversation, "materials.csv")
        self.runtime._conversation_file_selector_mode = "enforce_narrow"
        auto_response = await self.submit_message(
            conversation_id=auto_conversation,
            capability_id=None,
            content="请分析材料文件。",
            metadata={"file_requirement_profile": {"required": True}},
        )
        self.assertEqual(auto_response.status_code, 202, auto_response.text)
        auto_task_id = auto_response.json()["task_id"]
        await self.runtime._await_existing_execution(auto_task_id)

        def selector_generator(_prompt: str, **_kwargs) -> str:
            return "not-json SECRET_VALUE_SHOULD_NOT_BE_AUDITED"

        await self.reconfigure_runtime(skill_input_text_generator=selector_generator)
        self.runtime._conversation_file_selector_mode = "enforce_narrow"
        invalid_conversation = "conv-file-audit-catalog-invalid"
        first = await self._upload_csv(invalid_conversation, "first.csv")
        await self._upload_csv(invalid_conversation, "second.csv")
        invalid_response = await self.submit_message(
            conversation_id=invalid_conversation,
            capability_id=None,
            content="请分析材料文件。",
            metadata={
                "file_requirement_profile": {
                    "required": True,
                    "disambiguation_hint": "storage_key=/tmp/private/secret.txt",
                    "context_notes": ["safe note", "token=abc"],
                }
            },
        )
        self.assertEqual(invalid_response.status_code, 202, invalid_response.text)
        invalid_task_id = invalid_response.json()["task_id"]
        invalid_interrupt = next(item for item in await self.runtime.list_interrupts(invalid_task_id) if item["status"] == "open")
        answer = await self.answer_interrupt_with_chat(
            conversation_id=invalid_conversation,
            interrupt_id=invalid_interrupt["interrupt_id"],
            content=f"使用 {first}",
        )
        self.assertEqual(answer.status_code, 202, answer.text)
        await self.runtime._await_existing_execution(invalid_task_id)

        events = [
            *(await self.runtime.storage.list_events_for_task(auto_task_id)),
            *(await self.runtime.storage.list_events_for_task(invalid_task_id)),
        ]
        catalog_events = {
            "conversation_file.file_selector_invoked",
            "conversation_file.file_selector_decision_recorded",
            "conversation_file.file_selector_invalid_output",
            "conversation_file.file_selector_clarification_requested",
            "conversation_file.file_selector_resumed_from_interrupt",
            "conversation_file.file_selector_auto_bound",
        }
        by_type = {event.event_type: event for event in events if event.event_type in catalog_events}

        self.assertEqual(set(by_type), catalog_events)
        self.assertEqual(
            next(
                event
                for event in await self.runtime.storage.list_events_for_task(auto_task_id)
                if event.event_type == "conversation_file.file_selector_auto_bound"
            ).payload["selected_upload_ids"],
            [auto_upload],
        )
        for event in by_type.values():
            self.assertEqual(event.visibility, EventVisibility.AUDIT_ONLY)
            self._assert_audit_payload_redacted(event.payload)

    async def test_multiple_active_files_are_available_without_selector_interrupt(self) -> None:
        self.runtime._conversation_file_selector_mode = "enforce_narrow"
        conversation_id = "conv-file-ambiguous"
        await self._upload_csv(conversation_id, "materials.csv", "ped_id,value\nA001,1\n")
        await self._upload_csv(conversation_id, "materials.csv", "ped_id,value\nB001,2\n")

        response = await self.submit_message(
            conversation_id=conversation_id,
            capability_id=None,
            content="请用上传的文件做摘要。",
        )

        self.assertEqual(response.status_code, 202, response.text)
        task_id = response.json()["task_id"]
        interrupts = await self.runtime.list_interrupts(task_id)
        self.assertEqual(interrupts, [])
        await self.wait_for_terminal_task(task_id)
        attachments = await self.runtime.storage.list_task_input_attachments_for_task(task_id)
        self.assertEqual(attachments, [])
        resolved = await self.runtime.resolve_conversation_uploads_for_message(conversation_id, "acc-1")
        self.assertEqual(len(resolved["uploaded_artifacts"]), 2)

    async def test_file_requirement_profile_rejects_legacy_aliases(self) -> None:
        with self.assertRaisesRegex(FileRequirementProfileError, "accepted_file_types"):
            FileRequirementProfile.from_mapping({"required": True, "accepted_file_types": ["csv"]})

    async def test_request_metadata_allows_unrelated_intent_but_rejects_nested_legacy_fields(self) -> None:
        profile = self.runtime._file_requirement_profile_for_request(
            request=type("Request", (), {"metadata": {"intent": "qa"}, "content": "请解释概念"})(),
            metadata={"intent": "qa"},
        )
        self.assertFalse(profile.is_meaningful())

        with self.assertRaisesRegex(FileRequirementProfileError, "requires_file"):
            self.runtime._file_requirement_profile_for_request(
                request=type("Request", (), {"metadata": {"file_selection": {"requires_file": True}}, "content": "请处理文件"})(),
                metadata={"file_selection": {"requires_file": True}},
            )

    async def test_file_requirement_profile_accepts_final_fields(self) -> None:
        profile = FileRequirementProfile.from_mapping(
            {
                "source": "metadata",
                "required": True,
                "allow_multiple": False,
                "expected_content": ["材料表"],
                "supported_file_types": ["csv"],
                "helpful_columns": ["ped_id"],
                "disambiguation_hint": "优先材料表",
                "context_notes": ["metadata declared final file_selection"],
            }
        )

        self.assertTrue(profile.required)
        self.assertEqual(profile.supported_file_types, ("csv",))
        self.assertEqual(profile.source, "metadata")

    async def test_trigger_detector_ignores_ordinary_query_and_declines(self) -> None:
        detector = FileSelectionTriggerDetector()

        self.assertEqual(
            detector.should_trigger(
                text="请解释一下区组设计的基本概念。",
                profile=FileRequirementProfile(),
                has_explicit_uploads=False,
                active_file_count=2,
            ),
            (False, "ordinary_query"),
        )
        self.assertEqual(
            detector.should_trigger(
                text="这次不用上传的文件，只回答概念。",
                profile=FileRequirementProfile(required=True),
                has_explicit_uploads=False,
                active_file_count=2,
            ),
            (False, "explicit_or_declined"),
        )

    async def test_trigger_detector_continuation_requires_recent_usage_shadow(self) -> None:
        trigger, reason = FileSelectionTriggerDetector().should_trigger(
            text="继续用刚才那个数据。",
            profile=FileRequirementProfile(),
            has_explicit_uploads=False,
            active_file_count=2,
        )

        self.assertTrue(trigger)
        self.assertEqual(reason, "recent_usage_reference")

    async def test_enforce_narrow_ignores_non_file_ordinal_language(self) -> None:
        conversation_id = "conv-file-non-file-ordinal"
        await self._upload_csv(conversation_id, "materials.csv")
        resources = await self.runtime.storage.list_conversation_file_resources(
            conversation_id,
            "acc-1",
            include_deleted=False,
        )
        candidates = tuple(candidate_from_resource(resource) for resource in resources)

        self.assertEqual(
            FileSelectionTriggerDetector().should_trigger_enforce_narrow(
                text="请解释第一阶段的试验设计概念。",
                profile=FileRequirementProfile(),
                has_explicit_uploads=False,
                candidates=candidates,
            ),
            (False, "ordinary_query"),
        )

    async def test_enforce_narrow_does_not_bind_generic_ordinal_submit_message(self) -> None:
        self.runtime._conversation_file_selector_mode = "enforce_narrow"
        conversation_id = "conv-file-generic-ordinal"
        await self._upload_csv(conversation_id, "materials.csv")
        await self._upload_csv(conversation_id, "other.csv")

        response = await self.submit_message(
            conversation_id=conversation_id,
            capability_id=None,
            content="请先解释第一个问题的试验设计概念。",
        )

        self.assertEqual(response.status_code, 202, response.text)
        task_id = response.json()["task_id"]
        await self.runtime._await_existing_execution(task_id)
        interrupts = await self.runtime.list_interrupts(task_id)
        attachments = await self.runtime.storage.list_task_input_attachments_for_task(task_id)
        self.assertEqual(interrupts, [])
        self.assertEqual(attachments, [])

    async def test_selector_mode_normalization_rejects_legacy_enforce_alias(self) -> None:
        self.assertEqual(self.runtime._normalize_conversation_file_selector_mode("disabled"), "disabled")
        self.assertEqual(self.runtime._normalize_conversation_file_selector_mode("shadow"), "shadow")
        self.assertEqual(self.runtime._normalize_conversation_file_selector_mode("enforce_narrow"), "enforce_narrow")
        self.assertEqual(
            self.runtime._normalize_conversation_file_selector_mode("enforce_guarded_multi"),
            "enforce_guarded_multi",
        )
        self.assertEqual(self.runtime._normalize_conversation_file_selector_mode("enforce"), "disabled")
        self.assertEqual(self.runtime._normalize_conversation_file_selector_mode("unexpected"), "disabled")

    async def test_selector_mode_runtime_bootstrap_records_invalid_mode_config_error(self) -> None:
        with patch.dict("os.environ", {"MAF_CONVERSATION_FILE_SELECTOR_MODE": "enforce"}, clear=False):
            await self.reconfigure_runtime()

        self.assertEqual(self.runtime._conversation_file_selector_mode, "disabled")
        self.assertFalse(self.runtime._conversation_file_selector_guarded_multi_select)
        audit_log = (self.workspace / "audit.jsonl").read_text(encoding="utf-8")
        self.assertIn("conversation_file.file_selector_config_invalid", audit_log)
        self.assertIn("invalid_conversation_file_selector_mode", audit_log)
        self.assertIn("enforce", audit_log)
        self.assertNotIn("MAF_POSTGRES_STATE_DSN", audit_log)
        self.assertNotIn("content_base64", audit_log)

        with patch.dict("os.environ", {"MAF_CONVERSATION_FILE_SELECTOR_MODE": "enforce_guarded_multi"}, clear=False):
            await self.reconfigure_runtime()

        self.assertEqual(self.runtime._conversation_file_selector_mode, "enforce_guarded_multi")
        self.assertTrue(self.runtime._conversation_file_selector_guarded_multi_select)

        with patch.dict("os.environ", {"MAF_CONVERSATION_FILE_SELECTOR_MODE": "shadow"}, clear=False):
            await self.reconfigure_runtime()

        self.assertEqual(self.runtime._conversation_file_selector_mode, "shadow")
        self.assertFalse(self.runtime._conversation_file_selector_guarded_multi_select)

    async def test_disabled_rollback_keeps_context_without_selector_attachment_or_audit(self) -> None:
        selector_calls = 0
        captured_prompts: list[str] = []

        def selector_generator(_prompt: str, **_kwargs) -> str:
            nonlocal selector_calls
            selector_calls += 1
            return json.dumps({"decision": "select_one", "upload_ids": ["upl-deadbeef0000"], "confidence": 1})

        def main_agent_generator(prompt: str, **_kwargs) -> list[str]:
            captured_prompts.append(str(prompt))
            return ["完成。"]

        await self.reconfigure_runtime(
            main_agent_stream_generator=main_agent_generator,
            skill_input_text_generator=selector_generator,
        )
        self.runtime._conversation_file_selector_mode = "disabled"
        conversation_id = "conv-file-disabled-rollback"
        upload_id = await self._upload_csv(conversation_id, "materials.csv")

        response = await self.submit_message(
            conversation_id=conversation_id,
            capability_id=None,
            content="请用当前会话文件做一个摘要。",
            metadata={"file_requirement_profile": {"required": True}},
        )

        self.assertEqual(response.status_code, 202, response.text)
        task_id = response.json()["task_id"]
        await self.wait_for_terminal_task(task_id)
        self.assertEqual(selector_calls, 0)
        self.assertEqual(await self.runtime.storage.list_task_input_attachments_for_task(task_id), [])
        event_types = [event.event_type for event in await self.runtime.storage.list_events_for_task(task_id)]
        self.assertFalse(any(event_type.startswith("conversation_file.file_selector_") for event_type in event_types))
        self.assertTrue(captured_prompts)
        prompt = "\n".join(captured_prompts)
        self.assertIn("materials.csv", prompt)
        self.assertIn(upload_id, prompt)
        history = await self.client.get(f"/api/v1/conversations/{conversation_id}/messages")
        self.assertEqual(history.status_code, 200, history.text)
        self.assertTrue(
            any(message["message_type"] == "file_upload" and message["metadata"]["upload_id"] == upload_id for message in history.json()["messages"])
        )

    async def test_shadow_to_enforce_narrow_ordinary_question_keeps_execution_shape(self) -> None:
        async def run_mode(mode: str) -> tuple[list[str], list[str], list[str]]:
            prompts: list[str] = []

            def main_agent_generator(prompt: str, **_kwargs) -> list[str]:
                prompts.append(str(prompt))
                return ["完成。"]

            await self.reconfigure_runtime(main_agent_stream_generator=main_agent_generator)
            self.runtime._conversation_file_selector_mode = mode
            conversation_id = f"conv-file-ordinary-{mode}"
            await self._upload_csv(conversation_id, "materials.csv")
            response = await self.submit_message(
                conversation_id=conversation_id,
                capability_id=None,
                content="请解释一下随机区组设计的基本概念。",
            )
            self.assertEqual(response.status_code, 202, response.text)
            task_id = response.json()["task_id"]
            await self.wait_for_terminal_task(task_id)
            attachments = await self.runtime.storage.list_task_input_attachments_for_task(task_id)
            interrupts = await self.runtime.list_interrupts(task_id)
            event_types = [event.event_type for event in await self.runtime.storage.list_events_for_task(task_id)]
            self.assertEqual(attachments, [])
            self.assertEqual(interrupts, [])
            return prompts, event_types, [attachment.source_upload_id for attachment in attachments]

        shadow_prompts, shadow_events, shadow_attachment_ids = await run_mode("shadow")
        enforce_prompts, enforce_events, enforce_attachment_ids = await run_mode("enforce_narrow")

        self.assertEqual(shadow_attachment_ids, enforce_attachment_ids)
        self.assertFalse(any(event_type == "conversation_file.file_selector_auto_bound" for event_type in shadow_events))
        self.assertFalse(any(event_type == "conversation_file.file_selector_auto_bound" for event_type in enforce_events))
        self.assertTrue(shadow_prompts)
        self.assertTrue(enforce_prompts)

    async def test_upload_id_token_requires_exact_generated_shape(self) -> None:
        conversation_id = "conv-file-upload-id-token"
        await self._upload_csv(conversation_id, "materials.csv")
        resources = await self.runtime.storage.list_conversation_file_resources(
            conversation_id,
            "acc-1",
            include_deleted=False,
        )
        candidates = tuple(candidate_from_resource(resource) for resource in resources)

        embedded_existing = deterministic_file_decision(
            text=f"请使用 xxx{candidates[0].upload_id}yyy 这个文件。",
            profile=FileRequirementProfile(),
            candidates=candidates,
        )
        self.assertNotEqual(embedded_existing.reason_code, "explicit_upload_id")
        answer_embedded_existing = FileSelectionAnswerResolver().resolve(
            f"xxx{candidates[0].upload_id}yyy",
            candidates,
        )
        self.assertNotEqual(answer_embedded_existing.reason_code, "explicit_upload_id")

        embedded = deterministic_file_decision(
            text="请使用 xxxupl-abcdef123456yyy 这个文件。",
            profile=FileRequirementProfile(),
            candidates=candidates,
        )
        self.assertNotEqual(embedded.reason_code, "unknown_upload_id")

        exact_unknown = deterministic_file_decision(
            text="请使用 upl-abcdef123456 这个文件。",
            profile=FileRequirementProfile(),
            candidates=candidates,
        )
        self.assertEqual(exact_unknown.reason_code, "unknown_upload_id")

        mixed_existing_unknown = deterministic_file_decision(
            text=f"请同时使用 {candidates[0].upload_id} 和 upl-abcdef123456。",
            profile=FileRequirementProfile(),
            candidates=candidates,
        )
        self.assertEqual(mixed_existing_unknown.decision, "ambiguous")
        self.assertEqual(mixed_existing_unknown.reason_code, "unknown_upload_id")
        self.assertEqual(mixed_existing_unknown.upload_ids, ())
        answer_mixed_existing_unknown = FileSelectionAnswerResolver().resolve(
            f"使用 {candidates[0].upload_id} 和 upl-abcdef123456",
            candidates,
        )
        self.assertEqual(answer_mixed_existing_unknown.decision, "ambiguous")
        self.assertEqual(answer_mixed_existing_unknown.reason_code, "unknown_upload_id")
        self.assertEqual(answer_mixed_existing_unknown.upload_ids, ())

        trigger, reason = FileSelectionTriggerDetector().should_trigger(
            text="请使用 xxxupl-abcdef123456yyy 这个文件。",
            profile=FileRequirementProfile(),
            has_explicit_uploads=False,
            active_file_count=1,
        )
        self.assertTrue(trigger)
        self.assertEqual(reason, "query_reference")

        trigger, reason = FileSelectionTriggerDetector().should_trigger(
            text="请使用 upl-abcdef123456 这个文件。",
            profile=FileRequirementProfile(),
            has_explicit_uploads=False,
            active_file_count=1,
        )
        self.assertTrue(trigger)
        self.assertEqual(reason, "explicit_upload_id_reference")

    async def test_explicit_upload_ids_skip_shadow_selector(self) -> None:
        self.runtime._conversation_file_selector_mode = "shadow"
        conversation_id = "conv-file-explicit-skip"
        upload_id = await self._upload_csv(conversation_id, "materials.csv")

        response = await self.submit_message(
            conversation_id=conversation_id,
            capability_id=None,
            content="请用显式文件做摘要。",
            metadata={"upload_ids": [upload_id], "file_requirement_profile": {"required": True}},
        )

        self.assertEqual(response.status_code, 202, response.text)
        task_id = response.json()["task_id"]
        await self.runtime._await_existing_execution(task_id)
        event_types = [event.event_type for event in await self.runtime.storage.list_events_for_task(task_id)]
        self.assertNotIn("conversation_file.file_selector_invoked", event_types)

    async def test_required_file_with_no_active_files_opens_missing_file_clarification(self) -> None:
        self.runtime._conversation_file_selector_mode = "enforce_narrow"
        response = await self.submit_message(
            conversation_id="conv-file-required-missing",
            capability_id=None,
            content="请分析材料文件。",
            metadata={"file_requirement_profile": {"required": True, "expected_content": ["材料表"]}},
        )

        self.assertEqual(response.status_code, 202, response.text)
        task_id = response.json()["task_id"]
        interrupts = await self.runtime.list_interrupts(task_id)
        self.assertEqual(len(interrupts), 1)
        self.assertEqual(interrupts[0]["reason_code"], "file_selection_ambiguous")
        self.assertEqual(interrupts[0]["required_fields"]["_file_selection"]["reason_code"], "no_files_in_conversation")
        self.assertIn("还没有可用文件", interrupts[0]["question"])

    async def test_required_single_active_file_auto_binds_task_attachment(self) -> None:
        self.runtime._conversation_file_selector_mode = "enforce_narrow"
        conversation_id = "conv-file-required-auto-bind"
        upload_id = await self._upload_csv(conversation_id, "materials.csv")

        response = await self.submit_message(
            conversation_id=conversation_id,
            capability_id=None,
            content="请分析材料文件。",
            metadata={"file_requirement_profile": {"required": True, "expected_content": ["材料表"]}},
        )

        self.assertEqual(response.status_code, 202, response.text)
        task_id = response.json()["task_id"]
        await self.runtime._await_existing_execution(task_id)
        attachments = await self.runtime.storage.list_task_input_attachments_for_task(task_id)
        self.assertEqual(len(attachments), 1)
        self.assertEqual(attachments[0].source_kind, "file_selector")
        self.assertEqual(attachments[0].source_upload_id, upload_id)
        self.assertEqual(await self.runtime.list_interrupts(task_id), [])

    async def test_same_filename_required_opens_ambiguous_interrupt_without_binding(self) -> None:
        self.runtime._conversation_file_selector_mode = "enforce_narrow"
        conversation_id = "conv-file-required-same-name"
        await self._upload_csv(conversation_id, "materials.csv", "ped_id,value\nA001,1\n")
        await self._upload_csv(conversation_id, "materials.csv", "ped_id,value\nB001,2\n")

        response = await self.submit_message(
            conversation_id=conversation_id,
            capability_id=None,
            content="请分析 materials.csv。",
            metadata={"file_requirement_profile": {"required": True}},
        )

        self.assertEqual(response.status_code, 202, response.text)
        task_id = response.json()["task_id"]
        interrupts = await self.runtime.list_interrupts(task_id)
        self.assertEqual(len(interrupts), 1)
        self.assertEqual(interrupts[0]["reason_code"], "file_selection_ambiguous")
        self.assertEqual(len(interrupts[0]["required_fields"]["_file_selection"]["candidate_upload_ids"]), 2)
        self.assertEqual(await self.runtime.storage.list_task_input_attachments_for_task(task_id), [])

    async def test_same_filename_ambiguity_does_not_delegate_to_selector_guess(self) -> None:
        self.runtime._conversation_file_selector_mode = "enforce_narrow"
        conversation_id = "conv-file-same-name-no-selector-guess"
        first = await self._upload_csv(conversation_id, "materials.csv", "ped_id,value\nA001,1\n")
        second = await self._upload_csv(conversation_id, "materials.csv", "ped_id,value\nB001,2\n")
        selector_calls = 0

        def selector_generator(_prompt: str, **_kwargs) -> str:
            nonlocal selector_calls
            selector_calls += 1
            return json.dumps(
                {
                    "decision": "select_one",
                    "selected_upload_ids": [first],
                    "confidence": 0.99,
                    "reason_code": "guessed_duplicate_filename",
                }
            )

        await self.reconfigure_runtime(skill_input_text_generator=selector_generator)
        self.runtime._conversation_file_selector_mode = "enforce_narrow"

        response = await self.submit_message(
            conversation_id=conversation_id,
            capability_id=None,
            content="请分析 materials.csv。",
            metadata={"file_requirement_profile": {"required": True}},
        )

        self.assertEqual(response.status_code, 202, response.text)
        task_id = response.json()["task_id"]
        interrupts = await self.runtime.list_interrupts(task_id)
        self.assertEqual(selector_calls, 0)
        self.assertEqual(len(interrupts), 1)
        self.assertEqual(interrupts[0]["required_fields"]["_file_selection"]["reason_code"], "duplicate_filename_candidates")
        self.assertCountEqual(interrupts[0]["required_fields"]["_file_selection"]["candidate_upload_ids"], [first, second])
        self.assertEqual(await self.runtime.storage.list_task_input_attachments_for_task(task_id), [])

    async def test_body_upload_id_exact_token_binds_active_file(self) -> None:
        self.runtime._conversation_file_selector_mode = "enforce_narrow"
        conversation_id = "conv-file-body-upload-id"
        upload_id = await self._upload_csv(conversation_id, "materials.csv")

        response = await self.submit_message(
            conversation_id=conversation_id,
            capability_id=None,
            content=f"请使用 {upload_id} 这个文件做摘要。",
        )

        self.assertEqual(response.status_code, 202, response.text)
        task_id = response.json()["task_id"]
        await self.runtime._await_existing_execution(task_id)
        attachments = await self.runtime.storage.list_task_input_attachments_for_task(task_id)
        self.assertEqual(len(attachments), 1)
        self.assertEqual(attachments[0].source_kind, "file_selector")
        self.assertEqual(attachments[0].source_upload_id, upload_id)
        decision = next(
            event
            for event in await self.runtime.storage.list_events_for_task(task_id)
            if event.event_type == "conversation_file.file_selector_decision_recorded"
        )
        self.assertEqual(decision.payload["reason_code"], "explicit_upload_id")

    async def test_body_upload_id_invalid_tokens_fail_closed_before_llm_selector(self) -> None:
        selector_calls = 0

        def selector_generator(_prompt: str, **_kwargs) -> str:
            nonlocal selector_calls
            selector_calls += 1
            return json.dumps({"decision": "select_one", "upload_ids": ["upl-deadbeef0000"], "confidence": 1})

        await self.reconfigure_runtime(skill_input_text_generator=selector_generator)
        self.runtime._conversation_file_selector_mode = "enforce_narrow"

        active_conversation = "conv-file-unknown-body-id"
        active_id = await self._upload_csv(active_conversation, "active.csv")
        unknown = await self.submit_message(
            conversation_id=active_conversation,
            capability_id=None,
            content=f"请使用 {active_id} 和 upl-abcdef123456 这两个文件。",
        )
        self.assertEqual(unknown.status_code, 202, unknown.text)
        unknown_task = unknown.json()["task_id"]
        unknown_interrupts = await self.runtime.list_interrupts(unknown_task)
        self.assertEqual(len(unknown_interrupts), 1)
        self.assertEqual(unknown_interrupts[0]["required_fields"]["_file_selection"]["reason_code"], "unknown_upload_id")
        self.assertEqual(await self.runtime.storage.list_task_input_attachments_for_task(unknown_task), [])

        deleted_conversation = "conv-file-deleted-body-id"
        deleted_id = await self._upload_csv(deleted_conversation, "deleted.csv")
        deleted = await self.client.request(
            "DELETE",
            "/api/v1/conversations/uploads",
            json={"conversation_id": deleted_conversation, "upload_id": deleted_id},
        )
        self.assertEqual(deleted.status_code, 200, deleted.text)
        deleted_submit = await self.submit_message(
            conversation_id=deleted_conversation,
            capability_id=None,
            content=f"请使用 {deleted_id} 这个文件。",
        )
        self.assertEqual(deleted_submit.status_code, 202, deleted_submit.text)
        deleted_task = deleted_submit.json()["task_id"]
        deleted_interrupts = await self.runtime.list_interrupts(deleted_task)
        self.assertEqual(len(deleted_interrupts), 1)
        self.assertEqual(deleted_interrupts[0]["required_fields"]["_file_selection"]["reason_code"], "unknown_upload_id")

        foreign_id = await self._upload_csv("conv-file-foreign-source", "foreign.csv")
        foreign_submit = await self.submit_message(
            conversation_id="conv-file-foreign-target",
            capability_id=None,
            content=f"请使用 {foreign_id} 这个文件。",
        )
        self.assertEqual(foreign_submit.status_code, 202, foreign_submit.text)
        foreign_task = foreign_submit.json()["task_id"]
        foreign_interrupts = await self.runtime.list_interrupts(foreign_task)
        self.assertEqual(len(foreign_interrupts), 1)
        self.assertEqual(foreign_interrupts[0]["required_fields"]["_file_selection"]["reason_code"], "unknown_upload_id")
        self.assertEqual(selector_calls, 0)

    async def test_low_confidence_selector_opens_clarification_without_binding(self) -> None:
        conversation_id = "conv-file-low-confidence"
        first = await self._upload_csv(conversation_id, "first.csv")
        await self._upload_csv(conversation_id, "second.csv")

        def selector_generator(_prompt: str, **_kwargs) -> str:
            return json.dumps({"decision": "select_one", "selected_upload_ids": [first], "confidence": 0.5})

        await self.reconfigure_runtime(skill_input_text_generator=selector_generator)
        self.runtime._conversation_file_selector_mode = "enforce_narrow"

        response = await self.submit_message(
            conversation_id=conversation_id,
            capability_id=None,
            content="请分析材料文件。",
            metadata={"file_requirement_profile": {"required": True}},
        )

        self.assertEqual(response.status_code, 202, response.text)
        task_id = response.json()["task_id"]
        interrupts = await self.runtime.list_interrupts(task_id)
        self.assertEqual(len(interrupts), 1)
        self.assertEqual(interrupts[0]["required_fields"]["_file_selection"]["reason_code"], "low_confidence")
        self.assertEqual(await self.runtime.storage.list_task_input_attachments_for_task(task_id), [])

    async def test_invalid_selector_json_clarifies_and_records_audit_without_binding(self) -> None:
        def selector_generator(_prompt: str, **_kwargs) -> str:
            return "not-json SECRET_SHOULD_NOT_APPEAR"

        await self.reconfigure_runtime(skill_input_text_generator=selector_generator)
        self.runtime._conversation_file_selector_mode = "enforce_narrow"
        conversation_id = "conv-file-invalid-json-enforce"
        await self._upload_csv(conversation_id, "first.csv")
        await self._upload_csv(conversation_id, "second.csv")

        response = await self.submit_message(
            conversation_id=conversation_id,
            capability_id=None,
            content="请分析材料文件。",
            metadata={"file_requirement_profile": {"required": True}},
        )

        self.assertEqual(response.status_code, 202, response.text)
        task_id = response.json()["task_id"]
        interrupts = await self.runtime.list_interrupts(task_id)
        self.assertEqual(len(interrupts), 1)
        self.assertEqual(interrupts[0]["required_fields"]["_file_selection"]["reason_code"], "invalid_json")
        self.assertEqual(await self.runtime.storage.list_task_input_attachments_for_task(task_id), [])
        invalid = next(
            event
            for event in await self.runtime.storage.list_events_for_task(task_id)
            if event.event_type == "conversation_file.file_selector_invalid_output"
        )
        payload_text = json.dumps(invalid.payload, ensure_ascii=False)
        self.assertNotIn("SECRET_SHOULD_NOT_APPEAR", payload_text)

    async def test_invalid_selector_selected_ids_are_not_copied_to_audit(self) -> None:
        def selector_generator(_prompt: str, **_kwargs) -> str:
            return json.dumps(
                {
                    "decision": "select_one",
                    "selected_upload_ids": ["storage_key=/tmp/private/secret.txt"],
                    "confidence": 0.99,
                }
            )

        await self.reconfigure_runtime(skill_input_text_generator=selector_generator)
        self.runtime._conversation_file_selector_mode = "enforce_narrow"
        conversation_id = "conv-file-invalid-selector-id-audit"
        await self._upload_csv(conversation_id, "first.csv")
        await self._upload_csv(conversation_id, "second.csv")

        response = await self.submit_message(
            conversation_id=conversation_id,
            capability_id=None,
            content="请分析材料文件。",
            metadata={"file_requirement_profile": {"required": True}},
        )

        self.assertEqual(response.status_code, 202, response.text)
        task_id = response.json()["task_id"]
        decision = next(
            event
            for event in await self.runtime.storage.list_events_for_task(task_id)
            if event.event_type == "conversation_file.file_selector_decision_recorded"
        )
        payload_text = json.dumps(decision.payload, ensure_ascii=False)
        self.assertEqual(decision.payload["reason_code"], "unknown_upload_id")
        self.assertEqual(decision.payload["selected_upload_ids"], [])
        self.assertNotIn("storage_key", payload_text)
        self.assertNotIn("/tmp/private", payload_text)

    async def test_select_many_default_requires_confirmation_without_binding(self) -> None:
        conversation_id = "conv-file-select-many-confirm"
        first = await self._upload_csv(conversation_id, "first.csv")
        second = await self._upload_csv(conversation_id, "second.csv")

        def selector_generator(_prompt: str, **_kwargs) -> str:
            return json.dumps(
                {
                    "decision": "select_many",
                    "selected_upload_ids": [first, second],
                    "confidence": 0.95,
                    "reason_code": "comparison",
                }
            )

        await self.reconfigure_runtime(skill_input_text_generator=selector_generator)
        self.runtime._conversation_file_selector_mode = "enforce_narrow"

        response = await self.submit_message(
            conversation_id=conversation_id,
            capability_id=None,
            content="请比较材料文件。",
            metadata={"file_requirement_profile": {"required": True, "allow_multiple": True}},
        )

        self.assertEqual(response.status_code, 202, response.text)
        task_id = response.json()["task_id"]
        interrupts = await self.runtime.list_interrupts(task_id)
        self.assertEqual(len(interrupts), 1)
        self.assertEqual(interrupts[0]["required_fields"]["_file_selection"]["reason_code"], "multi_select_rollout_disabled")
        self.assertEqual(await self.runtime.storage.list_task_input_attachments_for_task(task_id), [])

    async def test_guarded_multi_auto_binds_clear_multi_intent(self) -> None:
        conversation_id = "conv-file-guarded-multi"
        first = await self._upload_csv(conversation_id, "first.csv")
        second = await self._upload_csv(conversation_id, "second.csv")

        def selector_generator(_prompt: str, **_kwargs) -> str:
            return json.dumps(
                {
                    "decision": "select_many",
                    "selected_upload_ids": [first, second],
                    "confidence": 0.96,
                    "reason_code": "comparison",
                }
            )

        await self.reconfigure_runtime(skill_input_text_generator=selector_generator)
        self.runtime._conversation_file_selector_mode = "enforce_guarded_multi"
        self.runtime._conversation_file_selector_guarded_multi_select = True

        response = await self.submit_message(
            conversation_id=conversation_id,
            capability_id=None,
            content="请比较 first.csv 和 second.csv 两个文件。",
        )

        self.assertEqual(response.status_code, 202, response.text)
        task_id = response.json()["task_id"]
        await self.runtime._await_existing_execution(task_id)
        attachments = await self.runtime.storage.list_task_input_attachments_for_task(task_id)
        self.assertEqual({attachment.source_upload_id for attachment in attachments}, {first, second})
        self.assertTrue(all(attachment.source_kind == "file_selector" for attachment in attachments))
        self.assertEqual(await self.runtime.list_interrupts(task_id), [])
        events = await self.runtime.storage.list_events_for_task(task_id)
        auto_bound = next(event for event in events if event.event_type == "conversation_file.file_selector_auto_bound")
        self.assertEqual(auto_bound.payload["multi_select_resolution"], "multi_select_auto_bound")
        self._assert_audit_payload_redacted(auto_bound.payload)

    async def test_guarded_multi_still_requires_explicit_multi_intent(self) -> None:
        conversation_id = "conv-file-guarded-multi-no-intent"
        first = await self._upload_csv(conversation_id, "first.csv")
        second = await self._upload_csv(conversation_id, "second.csv")

        def selector_generator(_prompt: str, **_kwargs) -> str:
            return json.dumps(
                {
                    "decision": "select_many",
                    "selected_upload_ids": [first, second],
                    "confidence": 0.96,
                    "reason_code": "comparison",
                }
            )

        await self.reconfigure_runtime(skill_input_text_generator=selector_generator)
        self.runtime._conversation_file_selector_mode = "enforce_guarded_multi"
        self.runtime._conversation_file_selector_guarded_multi_select = True

        response = await self.submit_message(
            conversation_id=conversation_id,
            capability_id=None,
            content="请比较第一个阶段和第二个阶段。",
            metadata={"file_requirement_profile": {"required": True}},
        )

        self.assertEqual(response.status_code, 202, response.text)
        task_id = response.json()["task_id"]
        interrupts = await self.runtime.list_interrupts(task_id)
        self.assertEqual(len(interrupts), 1)
        self.assertEqual(interrupts[0]["required_fields"]["_file_selection"]["reason_code"], "multi_select_requires_confirmation")
        self.assertEqual(await self.runtime.storage.list_task_input_attachments_for_task(task_id), [])

    async def test_guarded_multi_auto_binds_multiple_exact_upload_ids_with_multi_intent(self) -> None:
        conversation_id = "conv-file-guarded-multi-exact"
        first = await self._upload_csv(conversation_id, "first.csv")
        second = await self._upload_csv(conversation_id, "second.csv")
        self.runtime._conversation_file_selector_mode = "enforce_guarded_multi"
        self.runtime._conversation_file_selector_guarded_multi_select = True

        response = await self.submit_message(
            conversation_id=conversation_id,
            capability_id=None,
            content=f"请比较 {first} 和 {second} 两个文件。",
        )

        self.assertEqual(response.status_code, 202, response.text)
        task_id = response.json()["task_id"]
        await self.runtime._await_existing_execution(task_id)
        attachments = await self.runtime.storage.list_task_input_attachments_for_task(task_id)
        self.assertEqual({attachment.source_upload_id for attachment in attachments}, {first, second})
        self.assertTrue(all(attachment.source_kind == "file_selector" for attachment in attachments))
        self.assertEqual(await self.runtime.list_interrupts(task_id), [])

    async def test_guarded_multi_exact_upload_ids_without_multi_intent_clarifies(self) -> None:
        conversation_id = "conv-file-guarded-multi-exact-no-intent"
        first = await self._upload_csv(conversation_id, "first.csv")
        second = await self._upload_csv(conversation_id, "second.csv")
        self.runtime._conversation_file_selector_mode = "enforce_guarded_multi"
        self.runtime._conversation_file_selector_guarded_multi_select = True

        response = await self.submit_message(
            conversation_id=conversation_id,
            capability_id=None,
            content=f"请使用 {first} 和 {second}。",
        )

        self.assertEqual(response.status_code, 202, response.text)
        task_id = response.json()["task_id"]
        interrupts = await self.runtime.list_interrupts(task_id)
        self.assertEqual(len(interrupts), 1)
        self.assertEqual(interrupts[0]["required_fields"]["_file_selection"]["reason_code"], "multi_select_requires_confirmation")
        self.assertEqual(await self.runtime.storage.list_task_input_attachments_for_task(task_id), [])

    async def test_recent_usage_selects_previous_attachment_not_newer_upload(self) -> None:
        self.runtime._conversation_file_selector_mode = "enforce_narrow"
        conversation_id = "conv-file-recent-enforce"
        first = await self._upload_csv(conversation_id, "first.csv")
        await self.runtime.storage.save_task_input_attachment(
            TaskInputAttachment(
                attachment_id="att-recent-enforce",
                task_id="task-recent-enforce",
                conversation_id=conversation_id,
                source_kind="file_selector",
                source_upload_id=first,
                filename="first.csv",
                created_at=datetime(2026, 6, 19, 12, 0, 0),
                updated_at=datetime(2026, 6, 19, 12, 5, 0),
            )
        )
        await self._upload_csv(conversation_id, "second.csv")

        response = await self.submit_message(
            conversation_id=conversation_id,
            capability_id=None,
            content="继续用刚才那个数据做分析。",
        )

        self.assertEqual(response.status_code, 202, response.text)
        task_id = response.json()["task_id"]
        await self.runtime._await_existing_execution(task_id)
        attachments = await self.runtime.storage.list_task_input_attachments_for_task(task_id)
        self.assertEqual(len(attachments), 1)
        self.assertEqual(attachments[0].source_upload_id, first)
        self.assertEqual(attachments[0].source_kind, "file_selector")

    async def test_file_selection_interrupt_resume_binds_upload_id_answer(self) -> None:
        self.runtime._conversation_file_selector_mode = "enforce_narrow"
        conversation_id = "conv-file-interrupt-resume"
        first = await self._upload_csv(conversation_id, "materials.csv", "ped_id,value\nA001,1\n")
        await self._upload_csv(conversation_id, "materials.csv", "ped_id,value\nB001,2\n")
        submitted = await self.submit_message(
            conversation_id=conversation_id,
            capability_id=None,
            content="请分析 materials.csv。",
            metadata={"file_requirement_profile": {"required": True}},
        )
        self.assertEqual(submitted.status_code, 202, submitted.text)
        task_id = submitted.json()["task_id"]
        interrupt = next(item for item in await self.runtime.list_interrupts(task_id) if item["status"] == "open")

        answer = await self.answer_interrupt_with_chat(
            conversation_id=conversation_id,
            interrupt_id=interrupt["interrupt_id"],
            content=f"使用 {first}",
            client_message_id="client-file-selection-resume-replay",
        )

        self.assertEqual(answer.status_code, 202, answer.text)
        self.assertEqual(answer.json()["action"], "interrupt_resumed")
        await self.runtime._await_existing_execution(task_id)
        attachments = await self.runtime.storage.list_task_input_attachments_for_task(task_id)
        self.assertEqual(len(attachments), 1)
        self.assertEqual(attachments[0].source_upload_id, first)
        self.assertEqual(attachments[0].source_kind, "file_selector")
        self.assertIsNotNone(attachments[0].interrupt_answer_id)

        await self.runtime.storage.mark_conversation_file_resource_deleted(
            conversation_id,
            "acc-1",
            first,
            updated_at=datetime(2026, 8, 26, 12, 0, 0),
        )
        with patch.object(
            FileSelectionAnswerResolver,
            "resolve",
            side_effect=AssertionError("exact replay must not rerun file selector"),
        ):
            replay = await self.answer_interrupt_with_chat(
                conversation_id=conversation_id,
                interrupt_id=interrupt["interrupt_id"],
                content=f"使用 {first}",
                client_message_id="client-file-selection-resume-replay",
            )
        self.assertEqual(replay.status_code, 400, replay.text)
        self.assertIn("Task is terminal", replay.text)
        replayed_attachments = await self.runtime.storage.list_task_input_attachments_for_task(task_id)
        self.assertEqual(len(replayed_attachments), 1)

    async def test_file_selection_answered_before_schedule_retries_continuation_from_claim(self) -> None:
        self.runtime._conversation_file_selector_mode = "enforce_narrow"
        conversation_id = "conv-file-schedule-retry"
        first = await self._upload_csv(conversation_id, "materials.csv", "ped_id,value\nA001,1\n")
        await self._upload_csv(conversation_id, "materials.csv", "ped_id,value\nB001,2\n")
        submitted = await self.submit_message(
            conversation_id=conversation_id,
            capability_id=None,
            content="请分析 materials.csv。",
            metadata={"file_requirement_profile": {"required": True}},
        )
        task_id = submitted.json()["task_id"]
        interrupt = next(item for item in await self.runtime.list_interrupts(task_id) if item["status"] == "open")
        message_id = "client-file-schedule-retry"
        content = f"使用 {first}"
        answer_payload = {
            "client_request_id": message_id,
            "answer": {"text": content},
            "upload_ids": [],
            "file_selection_answer": content,
        }

        with patch.object(
            self.runtime,
            "_schedule_file_selection_agent_resume",
            side_effect=RuntimeError("schedule_failed"),
        ):
            with self.assertRaisesRegex(RuntimeError, "schedule_failed"):
                await self.runtime.answer_interrupt(
                    task_id,
                    interrupt["interrupt_id"],
                    answer_payload,
                    source_message_id=message_id,
                )

        await self.runtime.storage.mark_conversation_file_resource_deleted(
            conversation_id,
            "acc-1",
            first,
            updated_at=datetime(2026, 8, 26, 12, 30, 0),
        )

        scheduler = AsyncMock(return_value=None)
        with patch.object(
            self.runtime,
            "_schedule_file_selection_agent_resume",
            new=scheduler,
        ):
            replay = await self.runtime.answer_interrupt(
                task_id,
                interrupt["interrupt_id"],
                answer_payload,
                source_message_id=message_id,
            )
        self.assertEqual(replay["action"], "resumed")
        self.assertEqual(scheduler.await_count, 1)
        receipts = [
            event
            for event in await self.runtime.storage.list_events_for_task(task_id)
            if event.event_type == "task.interrupt_continuation_completed"
        ]
        self.assertEqual(len(receipts), 1)

    async def test_file_selection_interrupt_answer_mixed_valid_unknown_upload_id_stays_clarifying(self) -> None:
        self.runtime._conversation_file_selector_mode = "enforce_narrow"
        conversation_id = "conv-file-interrupt-mixed-id"
        first = await self._upload_csv(conversation_id, "materials.csv", "ped_id,value\nA001,1\n")
        await self._upload_csv(conversation_id, "materials.csv", "ped_id,value\nB001,2\n")
        submitted = await self.submit_message(
            conversation_id=conversation_id,
            capability_id=None,
            content="请分析 materials.csv。",
            metadata={"file_requirement_profile": {"required": True}},
        )
        task_id = submitted.json()["task_id"]
        interrupt = next(item for item in await self.runtime.list_interrupts(task_id) if item["status"] == "open")

        answer = await self.answer_interrupt_with_chat(
            conversation_id=conversation_id,
            interrupt_id=interrupt["interrupt_id"],
            content=f"使用 {first} 和 upl-abcdef123456",
        )

        self.assertEqual(answer.status_code, 202, answer.text)
        self.assertEqual(answer.json()["action"], "interrupt_clarification_answer")
        self.assertEqual(await self.runtime.storage.list_task_input_attachments_for_task(task_id), [])

    async def test_file_selection_clarification_exact_replay_keeps_one_user_message_and_one_audit(self) -> None:
        self.runtime._conversation_file_selector_mode = "enforce_narrow"
        conversation_id = "conv-file-interrupt-clarification-replay"
        first = await self._upload_csv(conversation_id, "materials.csv", "ped_id,value\nA001,1\n")
        await self._upload_csv(conversation_id, "materials.csv", "ped_id,value\nB001,2\n")
        submitted = await self.submit_message(
            conversation_id=conversation_id,
            capability_id=None,
            content="请分析 materials.csv。",
            metadata={"file_requirement_profile": {"required": True}},
        )
        task_id = submitted.json()["task_id"]
        interrupt = next(item for item in await self.runtime.list_interrupts(task_id) if item["status"] == "open")
        message_id = "client-file-clarification-replay"
        content = f"使用 {first} 和 upl-abcdef123456"

        first_answer = await self.answer_interrupt_with_chat(
            conversation_id=conversation_id,
            interrupt_id=interrupt["interrupt_id"],
            content=content,
            client_message_id=message_id,
        )
        self.assertEqual(first_answer.status_code, 202, first_answer.text)
        messages_before = await self.runtime.storage.list_messages_for_conversation(conversation_id)
        events_before = await self.runtime.storage.list_events_for_task(task_id)

        replay = await self.answer_interrupt_with_chat(
            conversation_id=conversation_id,
            interrupt_id=interrupt["interrupt_id"],
            content=content,
            client_message_id=message_id,
        )

        self.assertEqual(replay.status_code, 202, replay.text)
        self.assertEqual(replay.json()["action"], "interrupt_clarification_answer")
        messages_after = await self.runtime.storage.list_messages_for_conversation(conversation_id)
        events_after = await self.runtime.storage.list_events_for_task(task_id)
        self.assertEqual(len(messages_after), len(messages_before))
        replay_sensitive_types = {
            "conversation_file.file_selector_answer_claimed",
            "conversation_file.file_selector_invalid_output",
        }
        self.assertEqual(
            [event.event_type for event in events_after if event.event_type in replay_sensitive_types],
            [event.event_type for event in events_before if event.event_type in replay_sensitive_types],
        )
        self.assertEqual(
            [message.message_id for message in messages_after if message.role == MessageRole.USER].count(message_id),
            1,
        )
        user_message = await self.runtime.storage.get_message(message_id)
        assistant_message = await self.runtime.storage.get_message(
            f"{message_id}:file-selection-clarification"
        )
        self.assertIsNotNone(user_message)
        self.assertIsNotNone(assistant_message)
        self.assertEqual(assistant_message.created_at, user_message.created_at)
        self.assertEqual(await self.runtime.storage.list_interrupt_answers(interrupt["interrupt_id"]), [])

        changed_payload = await self.answer_interrupt_with_chat(
            conversation_id=conversation_id,
            interrupt_id=interrupt["interrupt_id"],
            content="改用另一个回答",
            client_message_id=message_id,
        )
        self.assertEqual(changed_payload.status_code, 409, changed_payload.text)

    async def test_guarded_multi_interrupt_answer_binds_multiple_exact_upload_ids_with_multi_intent(self) -> None:
        self.runtime._conversation_file_selector_mode = "enforce_guarded_multi"
        self.runtime._conversation_file_selector_guarded_multi_select = True
        conversation_id = "conv-file-interrupt-guarded-multi"
        first = await self._upload_csv(conversation_id, "materials.csv", "ped_id,value\nA001,1\n")
        second = await self._upload_csv(conversation_id, "materials.csv", "ped_id,value\nB001,2\n")
        submitted = await self.submit_message(
            conversation_id=conversation_id,
            capability_id=None,
            content="请分析 materials.csv。",
            metadata={"file_requirement_profile": {"required": True}},
        )
        self.assertEqual(submitted.status_code, 202, submitted.text)
        task_id = submitted.json()["task_id"]
        interrupt = next(item for item in await self.runtime.list_interrupts(task_id) if item["status"] == "open")

        answer = await self.answer_interrupt_with_chat(
            conversation_id=conversation_id,
            interrupt_id=interrupt["interrupt_id"],
            content=f"请比较 {first} 和 {second} 两个文件。",
        )

        self.assertEqual(answer.status_code, 202, answer.text)
        self.assertEqual(answer.json()["action"], "interrupt_resumed")
        await self.runtime._await_existing_execution(task_id)
        attachments = await self.runtime.storage.list_task_input_attachments_for_task(task_id)
        self.assertEqual({attachment.source_upload_id for attachment in attachments}, {first, second})
        self.assertTrue(all(attachment.source_kind == "file_selector" for attachment in attachments))
        self.assertTrue(all(attachment.interrupt_answer_id for attachment in attachments))
        events = await self.runtime.storage.list_events_for_task(task_id)
        resumed = [
            event
            for event in events
            if event.event_type == "conversation_file.file_selector_resumed_from_interrupt"
            and set(event.payload.get("selected_upload_ids", [])) == {first, second}
        ][-1]
        self.assertEqual(resumed.payload["multi_select_resolution"], "multi_select_confirmed_by_user")
        self._assert_audit_payload_redacted(resumed.payload)

    async def test_guarded_multi_interrupt_answer_exact_upload_ids_without_multi_intent_clarifies(self) -> None:
        self.runtime._conversation_file_selector_mode = "enforce_guarded_multi"
        self.runtime._conversation_file_selector_guarded_multi_select = True
        conversation_id = "conv-file-interrupt-guarded-multi-no-intent"
        first = await self._upload_csv(conversation_id, "materials.csv", "ped_id,value\nA001,1\n")
        second = await self._upload_csv(conversation_id, "materials.csv", "ped_id,value\nB001,2\n")
        submitted = await self.submit_message(
            conversation_id=conversation_id,
            capability_id=None,
            content="请分析 materials.csv。",
            metadata={"file_requirement_profile": {"required": True}},
        )
        task_id = submitted.json()["task_id"]
        interrupt = next(item for item in await self.runtime.list_interrupts(task_id) if item["status"] == "open")

        answer = await self.answer_interrupt_with_chat(
            conversation_id=conversation_id,
            interrupt_id=interrupt["interrupt_id"],
            content=f"使用 {first} 和 {second}。",
        )

        self.assertEqual(answer.status_code, 202, answer.text)
        self.assertEqual(answer.json()["action"], "interrupt_clarification_answer")
        self.assertEqual(await self.runtime.storage.list_task_input_attachments_for_task(task_id), [])

    async def test_enforce_narrow_interrupt_answer_allow_multiple_still_requires_guarded_mode(self) -> None:
        self.runtime._conversation_file_selector_mode = "enforce_narrow"
        conversation_id = "conv-file-interrupt-allow-multiple-narrow"
        first = await self._upload_csv(conversation_id, "first.csv", "ped_id,value\nA001,1\n")
        second = await self._upload_csv(conversation_id, "second.csv", "ped_id,value\nB001,2\n")
        submitted = await self.submit_message(
            conversation_id=conversation_id,
            capability_id=None,
            content="请分析材料文件。",
            metadata={"file_requirement_profile": {"required": True, "allow_multiple": True}},
        )
        self.assertEqual(submitted.status_code, 202, submitted.text)
        task_id = submitted.json()["task_id"]
        interrupt = next(item for item in await self.runtime.list_interrupts(task_id) if item["status"] == "open")

        answer = await self.answer_interrupt_with_chat(
            conversation_id=conversation_id,
            interrupt_id=interrupt["interrupt_id"],
            content=f"请比较 {first} 和 {second} 两个文件。",
        )

        self.assertEqual(answer.status_code, 202, answer.text)
        self.assertEqual(answer.json()["action"], "interrupt_clarification_answer")
        self.assertEqual(await self.runtime.storage.list_task_input_attachments_for_task(task_id), [])

    async def test_required_file_selection_interrupt_answer_cannot_skip_file(self) -> None:
        self.runtime._conversation_file_selector_mode = "enforce_narrow"
        conversation_id = "conv-file-required-no-skip"
        await self._upload_csv(conversation_id, "materials.csv", "ped_id,value\nA001,1\n")
        await self._upload_csv(conversation_id, "materials.csv", "ped_id,value\nB001,2\n")
        submitted = await self.submit_message(
            conversation_id=conversation_id,
            capability_id=None,
            content="请分析 materials.csv。",
            metadata={"file_requirement_profile": {"required": True}},
        )
        task_id = submitted.json()["task_id"]
        interrupt = next(item for item in await self.runtime.list_interrupts(task_id) if item["status"] == "open")

        answer = await self.answer_interrupt_with_chat(
            conversation_id=conversation_id,
            interrupt_id=interrupt["interrupt_id"],
            content="不用文件。",
        )

        self.assertEqual(answer.status_code, 202, answer.text)
        self.assertEqual(answer.json()["action"], "interrupt_clarification_answer")
        self.assertIn("assistant_message", answer.json())
        self.assertEqual(await self.runtime.storage.list_task_input_attachments_for_task(task_id), [])

    async def test_replacement_upload_in_file_selection_interrupt_keeps_interrupt_upload_provenance(self) -> None:
        self.runtime._conversation_file_selector_mode = "enforce_narrow"
        conversation_id = "conv-file-replacement-provenance"
        await self._upload_csv(conversation_id, "materials.csv", "ped_id,value\nA001,1\n")
        await self._upload_csv(conversation_id, "materials.csv", "ped_id,value\nB001,2\n")
        submitted = await self.submit_message(
            conversation_id=conversation_id,
            capability_id=None,
            content="请分析 materials.csv。",
            metadata={"file_requirement_profile": {"required": True}},
        )
        task_id = submitted.json()["task_id"]
        interrupt = next(item for item in await self.runtime.list_interrupts(task_id) if item["status"] == "open")
        replacement = await self._upload_csv(conversation_id, "replacement.csv", "ped_id,value\nC001,3\n")

        answer = await self.answer_interrupt_with_chat(
            conversation_id=conversation_id,
            interrupt_id=interrupt["interrupt_id"],
            content="用这个新上传的文件。",
            metadata={"upload_ids": [replacement]},
        )

        self.assertEqual(answer.status_code, 202, answer.text)
        await self.runtime._await_existing_execution(task_id)
        attachments = await self.runtime.storage.list_task_input_attachments_for_task(task_id)
        self.assertEqual(len(attachments), 1)
        self.assertEqual(attachments[0].source_upload_id, replacement)
        self.assertEqual(attachments[0].source_kind, "interrupt_answer_upload")

    async def test_replacement_upload_requiring_sheet_selection_preserves_interrupt_provenance_after_sheet_answer(self) -> None:
        self.runtime._conversation_file_selector_mode = "enforce_narrow"
        conversation_id = "conv-file-replacement-sheet-provenance"
        await self._upload_csv(conversation_id, "materials.csv", "ped_id,value\nA001,1\n")
        await self._upload_csv(conversation_id, "materials.csv", "ped_id,value\nB001,2\n")
        submitted = await self.submit_message(
            conversation_id=conversation_id,
            capability_id=None,
            content="请分析 materials.csv。",
            metadata={"file_requirement_profile": {"required": True}},
        )
        task_id = submitted.json()["task_id"]
        file_interrupt = next(item for item in await self.runtime.list_interrupts(task_id) if item["status"] == "open")
        replacement = await self._upload_multi_sheet_xlsx(conversation_id, "replacement.xlsx")

        replacement_answer = await self.answer_interrupt_with_chat(
            conversation_id=conversation_id,
            interrupt_id=file_interrupt["interrupt_id"],
            content="用这个新上传的表格。",
            metadata={"upload_ids": [replacement]},
        )

        self.assertEqual(replacement_answer.status_code, 202, replacement_answer.text)
        self.assertEqual(replacement_answer.json()["action"], "interrupt_sheet_selection_required")
        sheet_interrupt = next(
            item
            for item in await self.runtime.list_interrupts(task_id)
            if item["status"] == "open" and item["reason_code"] == "sheet_selection_required"
        )

        sheet_answer = await self.answer_interrupt_with_chat(
            conversation_id=conversation_id,
            interrupt_id=sheet_interrupt["interrupt_id"],
            content="选择 Beta",
            metadata={"upload_sheet_selections": {replacement: "Beta"}},
        )

        self.assertEqual(sheet_answer.status_code, 202, sheet_answer.text)
        await self.runtime._await_existing_execution(task_id)
        attachments = await self.runtime.storage.list_task_input_attachments_for_task(task_id)
        self.assertEqual(len(attachments), 1)
        self.assertEqual(attachments[0].source_upload_id, replacement)
        self.assertEqual(attachments[0].source_kind, "interrupt_answer_upload")
        self.assertEqual(attachments[0].selected_sheet, "Beta")

    async def test_selected_spreadsheet_opens_sheet_selection_then_binds_selector_sheet(self) -> None:
        self.runtime._conversation_file_selector_mode = "enforce_narrow"
        conversation_id = "conv-file-selector-sheet"
        upload_id = await self._upload_multi_sheet_xlsx(conversation_id, "materials.xlsx")

        submitted = await self.submit_message(
            conversation_id=conversation_id,
            capability_id=None,
            content="请分析材料文件。",
            metadata={"file_requirement_profile": {"required": True}},
        )

        self.assertEqual(submitted.status_code, 202, submitted.text)
        task_id = submitted.json()["task_id"]
        interrupt = next(item for item in await self.runtime.list_interrupts(task_id) if item["status"] == "open")
        self.assertEqual(interrupt["reason_code"], "sheet_selection_required")
        self.assertEqual(interrupt["required_fields"]["upload_sheet_selections"]["required_upload_ids"], [upload_id])

        answer = await self.answer_interrupt_with_chat(
            conversation_id=conversation_id,
            interrupt_id=interrupt["interrupt_id"],
            content="选择 Beta",
            metadata={"upload_sheet_selections": {upload_id: "Beta"}},
        )

        self.assertEqual(answer.status_code, 202, answer.text)
        await self.runtime._await_existing_execution(task_id)
        attachments = await self.runtime.storage.list_task_input_attachments_for_task(task_id)
        self.assertEqual(len(attachments), 1)
        self.assertEqual(attachments[0].source_upload_id, upload_id)
        self.assertEqual(attachments[0].source_kind, "file_selector")
        self.assertEqual(attachments[0].selected_sheet, "Beta")

    async def test_selector_prompt_excludes_text_sample_and_raw_preview_values(self) -> None:
        captured_prompts: list[str] = []

        def selector_generator(prompt: str, **_kwargs) -> str:
            captured_prompts.append(prompt)
            return json.dumps({"decision": "ambiguous", "confidence": 0.4, "reason_code": "test"})

        await self.reconfigure_runtime(skill_input_text_generator=selector_generator)
        self.runtime._conversation_file_selector_mode = "enforce_narrow"
        conversation_id = "conv-file-redaction"
        await self._upload_csv(conversation_id, "secret.txt", "SECRET_VALUE_XYZ\nsecond line\n")
        await self._upload_csv(conversation_id, "materials.csv", "ped_id,value\nA001,1\n")

        resources = await self.runtime.storage.list_conversation_file_resources(
            conversation_id,
            "acc-1",
            include_deleted=False,
        )
        candidates = tuple(candidate_from_resource(resource) for resource in resources)
        decision = await self.runtime._conversation_file_selection_decision(
            text="请用上传的文件做摘要。",
            profile=FileRequirementProfile(),
            candidates=candidates,
            metadata={},
        )

        self.assertEqual(decision.decision, "ambiguous")
        self.assertTrue(captured_prompts)
        prompt = captured_prompts[0]
        self.assertNotIn("SECRET_VALUE_XYZ", prompt)
        self.assertNotIn("content_base64", prompt)
        self.assertNotIn("storage_key", prompt)
        self.assertNotIn("mount_path", prompt)

    async def test_shadow_records_invalid_selector_output_without_raw_prompt(self) -> None:
        def selector_generator(_prompt: str, **_kwargs) -> str:
            return "not-json SECRET_VALUE_SHOULD_NOT_BE_AUDITED"

        await self.reconfigure_runtime(skill_input_text_generator=selector_generator)
        self.runtime._conversation_file_selector_mode = "shadow"
        conversation_id = "conv-file-invalid-selector"
        await self._upload_csv(conversation_id, "first.csv")
        await self._upload_csv(conversation_id, "second.csv")

        response = await self.submit_message(
            conversation_id=conversation_id,
            capability_id=None,
            content="请用上传的文件做摘要。",
        )

        self.assertEqual(response.status_code, 202, response.text)
        task_id = response.json()["task_id"]
        await self.runtime._await_existing_execution(task_id)
        events = await self.runtime.storage.list_events_for_task(task_id)
        invalid = next(event for event in events if event.event_type == "conversation_file.file_selector_invalid_output")
        payload_text = json.dumps(invalid.payload, ensure_ascii=False)
        self.assertEqual(invalid.payload["reason_code"], "invalid_json")
        self.assertNotIn("SECRET_VALUE_SHOULD_NOT_BE_AUDITED", payload_text)
        self.assertNotIn("user_message", payload_text)
        self.assertNotIn("content_base64", payload_text)
        self.assertNotIn("storage_key", payload_text)

    async def test_shadow_audit_sanitizes_profile_free_text(self) -> None:
        self.runtime._conversation_file_selector_mode = "shadow"
        conversation_id = "conv-file-audit-profile-redaction"
        await self._upload_csv(conversation_id, "materials.csv")

        response = await self.submit_message(
            conversation_id=conversation_id,
            capability_id=None,
            content="请用上传的文件做摘要。",
            metadata={
                "file_requirement_profile": {
                    "required": True,
                    "disambiguation_hint": "storage_key=/tmp/private/secret.txt",
                    "context_notes": [
                        "token=abc123",
                        "api_key=xyz",
                        "Authorization: Bearer abc",
                        "private_key=hidden",
                        "safe note",
                    ],
                    "expected_content": ["content_base64=abc", "材料表"],
                }
            },
        )

        self.assertEqual(response.status_code, 202, response.text)
        task_id = response.json()["task_id"]
        await self.runtime._await_existing_execution(task_id)
        decision = next(
            event
            for event in await self.runtime.storage.list_events_for_task(task_id)
            if event.event_type == "conversation_file.file_selector_decision_recorded"
        )
        payload_text = json.dumps(decision.payload, ensure_ascii=False)
        self.assertIn("safe note", payload_text)
        self.assertIn("材料表", payload_text)
        self.assertNotIn("storage_key", payload_text)
        self.assertNotIn("/tmp/private", payload_text)
        self.assertNotIn("token=abc123", payload_text)
        self.assertNotIn("api_key=xyz", payload_text)
        self.assertNotIn("Authorization", payload_text)
        self.assertNotIn("Bearer abc", payload_text)
        self.assertNotIn("private_key=hidden", payload_text)
        self.assertNotIn("content_base64=abc", payload_text)

    async def test_selector_candidates_exclude_deleted_resources(self) -> None:
        conversation_id = "conv-file-selector-deleted"
        active_id = await self._upload_csv(conversation_id, "active.csv")
        deleted_id = await self._upload_csv(conversation_id, "deleted.csv")
        deleted = await self.client.request(
            "DELETE",
            "/api/v1/conversations/uploads",
            json={"conversation_id": conversation_id, "upload_id": deleted_id},
        )
        self.assertEqual(deleted.status_code, 200, deleted.text)

        resources = await self.runtime.storage.list_conversation_file_resources(
            conversation_id,
            "acc-1",
            include_deleted=False,
        )
        candidates = tuple(candidate_from_resource(resource) for resource in resources)

        self.assertEqual([candidate.upload_id for candidate in candidates], [active_id])
        decision = deterministic_file_decision(
            text=f"请继续使用 {deleted_id} 这个文件。",
            profile=FileRequirementProfile(),
            candidates=candidates,
        )
        self.assertEqual(decision.decision, "ambiguous")
        self.assertEqual(decision.reason_code, "unknown_upload_id")

    async def test_repair_pending_disabled_context_uses_db_resources_not_stale_index(self) -> None:
        conversation_id = "conv-file-repair-disabled-db-truth"
        active_id = await self._upload_csv(conversation_id, "active.csv")
        deleted_id = await self._upload_csv(conversation_id, "deleted.csv")
        deleted = await self.client.request(
            "DELETE",
            "/api/v1/conversations/uploads",
            json={"conversation_id": conversation_id, "upload_id": deleted_id},
        )
        self.assertEqual(deleted.status_code, 200, deleted.text)
        (self.runtime.conversation_file_store.conversation_dir(conversation_id) / "index.md").write_text(
            f"# stale index\n- {deleted_id} deleted.csv should not be trusted\n",
            encoding="utf-8",
        )
        await self.runtime.storage.record_conversation_file_index_repair_required(
            conversation_id,
            reason_code="test_stale_index_pending",
            affected_upload_ids=(deleted_id,),
            now=self.runtime._utcnow_naive(),
        )
        self.runtime._conversation_file_selector_mode = "disabled"

        resolved = await self.runtime.resolve_conversation_uploads_for_message(conversation_id, "acc-1")
        response = await self.submit_message(
            conversation_id=conversation_id,
            capability_id=None,
            content="请用当前会话文件做摘要。",
        )

        self.assertEqual(response.status_code, 202, response.text)
        task_id = response.json()["task_id"]
        await self.runtime._await_existing_execution(task_id)
        self.assertEqual([artifact["upload_id"] for artifact in resolved["uploaded_artifacts"]], [active_id])
        serialized = json.dumps(resolved, ensure_ascii=False, default=str)
        self.assertNotIn(deleted_id, serialized)
        self.assertNotIn("deleted.csv", serialized)
        self.assertEqual(await self.runtime.storage.list_task_input_attachments_for_task(task_id), [])

    async def test_repair_pending_selector_candidates_use_db_resources_not_stale_index(self) -> None:
        conversation_id = "conv-file-repair-selector-db-truth"
        active_id = await self._upload_csv(conversation_id, "active.csv")
        deleted_id = await self._upload_csv(conversation_id, "deleted.csv")
        deleted = await self.client.request(
            "DELETE",
            "/api/v1/conversations/uploads",
            json={"conversation_id": conversation_id, "upload_id": deleted_id},
        )
        self.assertEqual(deleted.status_code, 200, deleted.text)
        (self.runtime.conversation_file_store.conversation_dir(conversation_id) / "index.md").write_text(
            f"# stale index\n- {deleted_id} deleted.csv should not be trusted\n",
            encoding="utf-8",
        )
        await self.runtime.storage.record_conversation_file_index_repair_required(
            conversation_id,
            reason_code="test_stale_index_pending",
            affected_upload_ids=(deleted_id,),
            now=self.runtime._utcnow_naive(),
        )
        self.runtime._conversation_file_selector_mode = "enforce_narrow"

        response = await self.submit_message(
            conversation_id=conversation_id,
            capability_id=None,
            content=f"请使用 {deleted_id} 这个文件。",
        )

        self.assertEqual(response.status_code, 202, response.text)
        task_id = response.json()["task_id"]
        interrupts = await self.runtime.list_interrupts(task_id)
        self.assertEqual(len(interrupts), 1)
        selection = interrupts[0]["required_fields"]["_file_selection"]
        self.assertEqual(selection["reason_code"], "unknown_upload_id")
        self.assertEqual(selection["candidate_upload_ids"], [active_id])
        serialized = json.dumps(selection, ensure_ascii=False, default=str)
        self.assertNotIn(deleted_id, serialized)
        self.assertNotIn("deleted.csv", serialized)
        self.assertEqual(await self.runtime.storage.list_task_input_attachments_for_task(task_id), [])

    async def test_recent_usage_ignores_file_upload_history_without_task_attachment(self) -> None:
        conversation_id = "conv-file-recent-history"
        first_id = await self._upload_csv(conversation_id, "first.csv")
        second_id = await self._upload_csv(conversation_id, "second.csv")
        resources = await self.runtime.storage.list_conversation_file_resources(
            conversation_id,
            "acc-1",
            include_deleted=False,
        )
        candidates = tuple(candidate_from_resource(resource, recent_usage=build_recent_usage(()).get(resource.file_id)) for resource in resources)

        self.assertEqual({candidate.upload_id for candidate in candidates}, {first_id, second_id})
        self.assertTrue(all(candidate.recent_usage is None for candidate in candidates))
        no_provenance = deterministic_file_decision(
            text="继续用刚才的文件。",
            profile=FileRequirementProfile(),
            candidates=candidates,
        )
        self.assertEqual(no_provenance.decision, "ambiguous")
        self.assertEqual(no_provenance.reason_code, "multiple_candidates")

        used_at = datetime(2026, 6, 19, 12, 0, 0)
        recent_usage = build_recent_usage(
            (
                TaskInputAttachment(
                    attachment_id="att-recent",
                    task_id="task-recent",
                    conversation_id=conversation_id,
                    source_kind="file_selector",
                    source_upload_id=second_id,
                    filename="second.csv",
                    created_at=used_at - timedelta(minutes=1),
                    updated_at=used_at,
                ),
            )
        )
        candidates_with_usage = tuple(candidate_from_resource(resource, recent_usage=recent_usage.get(resource.file_id)) for resource in resources)
        with_provenance = deterministic_file_decision(
            text="继续用刚才的文件。",
            profile=FileRequirementProfile(),
            candidates=candidates_with_usage,
        )

        self.assertEqual(with_provenance.decision, "select_one")
        self.assertEqual(with_provenance.upload_ids, (second_id,))
        self.assertEqual(with_provenance.reason_code, "recent_usage")

    async def test_required_file_skill_metadata_opens_selector_without_uploads(self) -> None:
        capability_id = self._write_file_required_skill()
        skill_root = self.workspace / "file-required-skill"
        await self.reconfigure_runtime(skill_roots=(skill_root,), public_skill_roots=(skill_root,))
        self.runtime._conversation_file_selector_mode = "enforce_narrow"

        response = await self.submit_message(
            conversation_id="conv-file-required-skill",
            capability_id=capability_id,
            content="执行这个 Skill。",
        )

        self.assertEqual(response.status_code, 202, response.text)
        task_id = response.json()["task_id"]
        interrupts = await self.runtime.list_interrupts(task_id)
        self.assertEqual(len(interrupts), 1)
        required = interrupts[0]["required_fields"]["_file_selection"]
        self.assertEqual(required["reason_code"], "no_files_in_conversation")
        self.assertTrue(required["profile"]["required"])
        self.assertIn("材料表", required["profile"]["expected_content"])
